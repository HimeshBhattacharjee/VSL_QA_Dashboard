import argparse
import io
import logging
import os
import re
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import PurePosixPath
from typing import Any, Dict, Iterable, List, Optional, Tuple

from botocore.exceptions import ClientError
from openpyxl import load_workbook

from models.calibration_data_models import (
    ensure_calibration_indexes,
    load_calibration_file_tracking,
    mark_calibration_file_processed,
    normalize_number,
    normalize_shift,
    parse_date_value,
    upsert_calibration_document,
)
from paths import get_qc_data_key
from s3_service import S3Service


LOGGER_NAME = "calibration_extractor"
logger = logging.getLogger(LOGGER_NAME)
if not logger.handlers:
    logging.basicConfig(
        level=os.getenv("CALIBRATION_LOG_LEVEL", "INFO"),
        format="%(asctime)s %(levelname)s %(name)s %(message)s",
    )

EXCEL_EXTENSIONS = (".xlsx", ".xlsm")
CALIBRATION_PREFIX_NAME = "Calibration Data"
MAX_HEADER_SCAN_ROWS = 20

HEADER_ALIASES = {
    "date": {"date"},
    "shift": {"shift"},
    "module_no": {"moduleno", "modulenumber", "moduleno."},
    "time": {"time"},
    "pmax": {"pmax", "pmaxw"},
    "voc": {"voc", "vocv"},
    "isc": {"isc", "isca"},
    "vmp": {"vmp", "vmpv"},
    "imp": {"imp", "impa"},
    "fill_factor": {"fillfactor", "ff"},
    "module_temp": {
        "temperature",
        "temperaturec",
        "moduletemp",
        "moduletempc",
        "moduletemperature",
        "moduletemperaturec",
    },
    "room_temp": {"roomtemp", "roomtempc", "roomtemperature", "roomtemperaturec"},
}
REQUIRED_HEADER_FIELDS = {"date", "shift", "module_no"}


@dataclass(frozen=True)
class S3CalibrationFile:
    key: str
    file_name: str
    etag: str
    last_modified: datetime
    size: int
    allowed_lines: Tuple[int, ...]


def _utc_naive(value: datetime) -> datetime:
    if value.tzinfo is not None:
        return value.astimezone(timezone.utc).replace(tzinfo=None)
    return value


def _clean_etag(etag: Any) -> str:
    return str(etag or "").strip('"')


def _is_excel_key(key: str) -> bool:
    file_name = PurePosixPath(key).name
    if file_name.startswith("~$"):
        return False
    return key.lower().endswith(EXCEL_EXTENSIONS)


def _normalized_label(value: Any) -> str:
    text = str(value or "").strip().lower()
    text = text.replace("\u00b0", "")
    return re.sub(r"[^a-z0-9]", "", text)


def _header_field(value: Any) -> Optional[str]:
    label = _normalized_label(value)
    if not label:
        return None
    for field, aliases in HEADER_ALIASES.items():
        if label in aliases:
            return field
    return None


def _allowed_lines_from_key(key: str) -> Tuple[int, ...]:
    parts = [part for part in key.split("/") if part]
    for part in parts:
        if re.search(r"\bLINE\s*[-_ ]?\s*II\b", part, re.IGNORECASE):
            return (3, 4)
    for part in parts:
        if re.search(r"\bLINE\s*[-_ ]?\s*I\b", part, re.IGNORECASE):
            return (1, 2)
    return ()


def _line_number_from_sheet(sheet_name: str) -> Optional[int]:
    compact = re.sub(r"[^A-Z0-9]", "", str(sheet_name or "").upper())
    match = re.fullmatch(r"LINE([1-4])", compact)
    if match:
        return int(match.group(1))
    match = re.search(r"LINE([1-4])(?!\d)", compact)
    if match:
        return int(match.group(1))
    return None


def _find_header_mapping(worksheet) -> Tuple[Optional[int], Dict[str, int]]:
    for row_number, row in enumerate(
        worksheet.iter_rows(min_row=1, max_row=MAX_HEADER_SCAN_ROWS, values_only=True),
        start=1,
    ):
        mapping: Dict[str, int] = {}
        for index, value in enumerate(row):
            field = _header_field(value)
            if field and field not in mapping:
                mapping[field] = index
        if REQUIRED_HEADER_FIELDS.issubset(mapping):
            return row_number, mapping
    return None, {}


def _row_has_data(row: Iterable[Any]) -> bool:
    return any(str(cell).strip() for cell in row if cell is not None)


def _value_from_mapping(row: Tuple[Any, ...], mapping: Dict[str, int], field: str) -> Any:
    index = mapping.get(field)
    if index is None or index >= len(row):
        return None
    return row[index]


def _contains_off(value: Any) -> bool:
    return bool(re.search(r"\bOFF\b", str(value or ""), re.IGNORECASE))


def _off_at_row_beginning(row: Tuple[Any, ...]) -> bool:
    for value in row[:5]:
        if value is None or str(value).strip() == "":
            continue
        return bool(re.match(r"^\s*OFF\b", str(value), re.IGNORECASE))
    return False


def _is_repeated_header_row(row: Tuple[Any, ...]) -> bool:
    fields = {_header_field(value) for value in row if value is not None}
    return len(fields.intersection(REQUIRED_HEADER_FIELDS)) >= 2


def _build_record(
    *,
    row: Tuple[Any, ...],
    mapping: Dict[str, int],
    line_number: int,
    source_file: str,
    source_sheet: str,
    row_number: int,
    current_date: Optional[str],
    current_shift: Optional[str],
) -> Tuple[Optional[Dict[str, Any]], Optional[str], Optional[str]]:
    date_value = _value_from_mapping(row, mapping, "date")
    shift_value = _value_from_mapping(row, mapping, "shift")

    parsed_date = parse_date_value(date_value) or current_date
    parsed_shift = normalize_shift(shift_value) or current_shift

    module_no = _value_from_mapping(row, mapping, "module_no")
    status = "OFF" if _contains_off(module_no) or _off_at_row_beginning(row) else "ACTIVE"

    current_date = parsed_date or current_date
    current_shift = parsed_shift or current_shift

    if not parsed_date or not parsed_shift:
        return None, current_date, current_shift

    has_measurement = any(
        normalize_number(_value_from_mapping(row, mapping, field)) is not None
        for field in ("pmax", "voc", "isc", "vmp", "imp", "fill_factor", "module_temp", "room_temp")
    )
    time_value = _value_from_mapping(row, mapping, "time")
    if status != "OFF" and not module_no and not time_value and not has_measurement:
        return None, current_date, current_shift

    record: Dict[str, Any] = {
        "date": parsed_date,
        "shift": parsed_shift,
        "line_number": line_number,
        "status": status,
        "source_file": source_file,
        "source_sheet": source_sheet,
        "row_number": row_number,
    }

    if status == "OFF":
        return record, current_date, current_shift

    record.update(
        {
            "time": time_value,
            "module_no": module_no,
            "pmax": _value_from_mapping(row, mapping, "pmax"),
            "voc": _value_from_mapping(row, mapping, "voc"),
            "isc": _value_from_mapping(row, mapping, "isc"),
            "vmp": _value_from_mapping(row, mapping, "vmp"),
            "imp": _value_from_mapping(row, mapping, "imp"),
            "fill_factor": _value_from_mapping(row, mapping, "fill_factor"),
            "module_temp": _value_from_mapping(row, mapping, "module_temp"),
            "room_temp": _value_from_mapping(row, mapping, "room_temp"),
        }
    )
    return record, current_date, current_shift


def extract_records_from_workbook(content: bytes, file_info: S3CalibrationFile) -> List[Dict[str, Any]]:
    workbook = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    records: List[Dict[str, Any]] = []

    try:
        for worksheet in workbook.worksheets:
            line_number = _line_number_from_sheet(worksheet.title)
            if line_number is None or line_number not in file_info.allowed_lines:
                continue

            logger.info("worksheet_detected file=%s sheet=%s line=%s", file_info.key, worksheet.title, line_number)
            header_row, mapping = _find_header_mapping(worksheet)
            if not header_row:
                logger.error("calibration_header_not_found file=%s sheet=%s", file_info.key, worksheet.title)
                continue

            current_date: Optional[str] = None
            current_shift: Optional[str] = None
            sheet_count = 0

            for row_number, row in enumerate(
                worksheet.iter_rows(min_row=header_row + 1, values_only=True),
                start=header_row + 1,
            ):
                if not _row_has_data(row) or _is_repeated_header_row(row):
                    continue
                record, current_date, current_shift = _build_record(
                    row=row,
                    mapping=mapping,
                    line_number=line_number,
                    source_file=file_info.key,
                    source_sheet=worksheet.title,
                    row_number=row_number,
                    current_date=current_date,
                    current_shift=current_shift,
                )
                if not record:
                    continue
                records.append(record)
                sheet_count += 1

            logger.info(
                "records_extracted file=%s sheet=%s line=%s records=%s",
                file_info.key,
                worksheet.title,
                line_number,
                sheet_count,
            )
    finally:
        workbook.close()

    return records


def list_s3_calibration_files(s3_prefix: str, s3_service: Optional[S3Service] = None) -> List[S3CalibrationFile]:
    s3_service = s3_service or S3Service()
    normalized_prefix = s3_prefix.rstrip("/") + "/"
    files: List[S3CalibrationFile] = []

    logger.info("calibration_extractor_started prefix=%s", normalized_prefix)
    paginator = s3_service.s3_client.get_paginator("list_objects_v2")
    for page in paginator.paginate(Bucket=s3_service.bucket_name, Prefix=normalized_prefix):
        for obj in page.get("Contents", []):
            key = obj.get("Key")
            if not key or key.endswith("/") or not _is_excel_key(key):
                continue
            allowed_lines = _allowed_lines_from_key(key)
            if not allowed_lines:
                logger.info("calibration_file_skipped_unknown_line key=%s", key)
                continue
            file_info = S3CalibrationFile(
                key=key,
                file_name=PurePosixPath(key).name,
                etag=_clean_etag(obj.get("ETag")),
                last_modified=_utc_naive(obj["LastModified"]),
                size=int(obj.get("Size", 0)),
                allowed_lines=allowed_lines,
            )
            files.append(file_info)
            logger.info(
                "file_discovered key=%s allowed_lines=%s etag=%s",
                file_info.key,
                file_info.allowed_lines,
                file_info.etag,
            )
    return files


def file_has_changed(file_info: S3CalibrationFile, tracking: Optional[Dict[str, Any]]) -> bool:
    if not tracking:
        return True
    existing_etag = _clean_etag(tracking.get("etag"))
    existing_modified = tracking.get("last_modified")
    if isinstance(existing_modified, datetime):
        existing_modified = _utc_naive(existing_modified)
    return existing_etag != file_info.etag or existing_modified != file_info.last_modified


def _group_records(records: List[Dict[str, Any]]) -> Dict[Tuple[str, str, int], List[Dict[str, Any]]]:
    grouped: Dict[Tuple[str, str, int], List[Dict[str, Any]]] = {}
    for record in records:
        date = parse_date_value(record.get("date"))
        shift = normalize_shift(record.get("shift"))
        line_number = int(record["line_number"])
        if not date or not shift:
            continue
        grouped.setdefault((date, shift, line_number), []).append(record)
    return grouped


def process_calibration_file(file_info: S3CalibrationFile, s3_service: S3Service) -> Dict[str, int]:
    summary = {"records_extracted": 0, "inserted": 0, "updated": 0, "skipped": 0, "errors": 0}
    try:
        response = s3_service.s3_client.get_object(Bucket=s3_service.bucket_name, Key=file_info.key)
        records = extract_records_from_workbook(response["Body"].read(), file_info)
        grouped = _group_records(records)
        summary["records_extracted"] = len(records)

        groups_upserted = 0
        for group_records in grouped.values():
            try:
                result = upsert_calibration_document(group_records)
                groups_upserted += 1
                if result.get("inserted"):
                    summary["inserted"] += 1
                elif result.get("updated"):
                    summary["updated"] += 1
            except Exception as exc:
                summary["errors"] += 1
                logger.exception(
                    "calibration_group_upsert_failed file=%s group_size=%s error=%s",
                    file_info.key,
                    len(group_records),
                    exc,
                )

        mark_calibration_file_processed(
            file_path=file_info.key,
            etag=file_info.etag,
            last_modified=file_info.last_modified,
            file_size=file_info.size,
            status="processed" if summary["errors"] == 0 else "processed_with_errors",
            records_extracted=len(records),
            groups_upserted=groups_upserted,
        )
        logger.info(
            "calibration_file_processed file=%s records=%s groups=%s summary=%s",
            file_info.key,
            len(records),
            groups_upserted,
            summary,
        )
        return summary
    except Exception as exc:
        summary["errors"] += 1
        mark_calibration_file_processed(
            file_path=file_info.key,
            etag=file_info.etag,
            last_modified=file_info.last_modified,
            file_size=file_info.size,
            status="error",
            records_extracted=summary["records_extracted"],
            error=str(exc),
        )
        logger.exception("calibration_file_processing_failed file=%s error=%s", file_info.key, exc)
        return summary


def process_s3_prefix(s3_prefix: str, *, force: bool = False, s3_service: Optional[S3Service] = None) -> Dict[str, int]:
    ensure_calibration_indexes()
    s3_service = s3_service or S3Service()
    summary = {
        "files_discovered": 0,
        "files_processed": 0,
        "files_skipped": 0,
        "records_extracted": 0,
        "inserted": 0,
        "updated": 0,
        "skipped": 0,
        "errors": 0,
    }

    try:
        files = list_s3_calibration_files(s3_prefix, s3_service)
        summary["files_discovered"] = len(files)
        tracking = load_calibration_file_tracking(file_info.key for file_info in files)

        for file_info in files:
            if not force and not file_has_changed(file_info, tracking.get(file_info.key)):
                summary["files_skipped"] += 1
                summary["skipped"] += 1
                logger.info("skipped_unchanged_file key=%s", file_info.key)
                continue

            result = process_calibration_file(file_info, s3_service)
            summary["files_processed"] += 1
            for key in ("records_extracted", "inserted", "updated", "skipped", "errors"):
                summary[key] += result.get(key, 0)

        logger.info("calibration_extraction_completed prefix=%s summary=%s", s3_prefix, summary)
        return summary
    except ClientError as exc:
        summary["errors"] += 1
        logger.exception("calibration_s3_extraction_failed prefix=%s error=%s", s3_prefix, exc)
        return summary
    except Exception as exc:
        summary["errors"] += 1
        logger.exception("calibration_extraction_failed prefix=%s error=%s", s3_prefix, exc)
        return summary


def main(force: bool = False) -> Dict[str, int]:
    return process_s3_prefix(get_qc_data_key(CALIBRATION_PREFIX_NAME), force=force)


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Extract Calibration Data reports from S3 into MongoDB")
    parser.add_argument("--force", action="store_true", help="Reprocess all discovered calibration Excel files")
    args = parser.parse_args()
    main(force=args.force)
