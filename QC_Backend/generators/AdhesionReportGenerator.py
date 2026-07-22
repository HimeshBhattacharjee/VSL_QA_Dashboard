from logging_utils import log_progress
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from copy import copy
import io
from paths import get_template_key, download_from_s3

FRONT_ADHESION_THRESHOLD = 60
BACK_ADHESION_THRESHOLD = 40
ADHESION_FAIL_FILL = PatternFill(start_color='FECACA', end_color='FECACA', fill_type='solid')
ADHESION_FAIL_FONT_COLOR = '9C0006'


def parse_adhesion_numeric_value(value):
    if value in (None, '', '-'):
        return None

    try:
        return float(str(value).strip())
    except (TypeError, ValueError):
        return None


def calculate_adhesion_average(form_data, indexes):
    numeric_values = []

    for index in indexes:
        numeric_value = parse_adhesion_numeric_value(form_data.get(f'adhesion_data_{index}'))
        if numeric_value is not None:
            numeric_values.append(numeric_value)

    if not numeric_values:
        return '0.00'

    return f"{sum(numeric_values) / len(numeric_values):.2f}"


def build_adhesion_averages(form_data):
    return {
        'frontMinAvg': calculate_adhesion_average(form_data, [0, 4, 8, 12, 16]),
        'frontMaxAvg': calculate_adhesion_average(form_data, [1, 5, 9, 13, 17]),
        'backMinAvg': calculate_adhesion_average(form_data, [2, 6, 10, 14, 18]),
        'backMaxAvg': calculate_adhesion_average(form_data, [3, 7, 11, 15, 19]),
    }


def get_adhesion_threshold(field_name):
    try:
        field_index = int(field_name.rsplit('_', 1)[1])
    except (IndexError, ValueError):
        return None

    return FRONT_ADHESION_THRESHOLD if field_index % 4 in (0, 1) else BACK_ADHESION_THRESHOLD


def apply_adhesion_fail_style(cell):
    cell.fill = ADHESION_FAIL_FILL
    updated_font = copy(cell.font) if cell.font else Font()
    updated_font.color = ADHESION_FAIL_FONT_COLOR
    cell.font = updated_font

def fill_adhesion_basic_info(worksheet, adhesion_data):
    try:
        form_data = adhesion_data.get('form_data', {})
        basic_info_mapping = {
            'adhesion_editable_0': 'E5',
            'adhesion_editable_1': 'C8',
            'adhesion_editable_2': 'C9',
            'adhesion_editable_3': 'C10',
            'adhesion_editable_4': 'C15',
            'adhesion_editable_5': 'D15',
            'adhesion_editable_6': 'E15',
            'adhesion_editable_19': 'C21',
            'adhesion_editable_20': 'C22',
            'adhesion_editable_21': 'C23',
            'adhesion_editable_22': 'C24',
            'adhesion_editable_23': 'C25',
            'adhesion_editable_24': 'C26',
            'adhesion_editable_25': 'C27',
            'preparedBySignature': 'B38',
            'verifiedBySignature': 'E38'
        }
        
        # Fill basic info fields
        for field, cell_ref in basic_info_mapping.items():
            if field in form_data and form_data[field]:
                worksheet[cell_ref] = form_data[field]
        
        # Handle date field separately
        if 'testDate' in form_data and form_data['testDate']:
            worksheet['C6'] = form_data['testDate']
            log_progress(f"Date filled: {form_data['testDate']}")
        
        # Handle shift field separately
        if 'shift' in form_data and form_data['shift']:
            worksheet['C7'] = form_data['shift']
            log_progress(f"Shift filled: {form_data['shift']}")
        if form_data.get('lineNumber'):
            worksheet['C12'] = form_data['lineNumber']
        
        # Handle laminator field separately
        if 'laminator' in form_data and form_data['laminator']:
            worksheet['C11'] = form_data['laminator']
            log_progress(f"Laminator filled: {form_data['laminator']}")
        
        # Handle lamination position field separately
        if 'laminationPosition' in form_data and form_data['laminationPosition']:
            worksheet['C12'] = form_data['laminationPosition']
            log_progress(f"Lamination Position filled: {form_data['laminationPosition']}")
        
        # Handle lamination parameters from JSON
        if 'lamParams' in form_data and form_data['lamParams']:
            import json
            lam_params = json.loads(form_data['lamParams'])
            
            # Mapping for lamination parameters to Excel cells
            # Adjust these cell references based on your actual Excel template
            lam_params_mapping = {
                # Lam1 parameters
                'lam1_pumpingTime': 'C16',    # Adjust cell reference as needed
                'lam1_pressingTime': 'C17',   # Adjust cell reference as needed
                'lam1_ventingTime': 'C18',    # Adjust cell reference as needed
                'lam1_processTime': 'C19',    # Adjust cell reference as needed
                
                # Lam2 parameters
                'lam2_pumpingTime': 'D16',    # Adjust cell reference as needed
                'lam2_pressingTime': 'D17',   # Adjust cell reference as needed
                'lam2_ventingTime': 'D18',    # Adjust cell reference as needed
                'lam2_processTime': 'D19',    # Adjust cell reference as needed
                
                # Lam3 parameters
                'lam3_pumpingTime': 'E16',    # Adjust cell reference as needed
                'lam3_pressingTime': 'E17',   # Adjust cell reference as needed
                'lam3_ventingTime': 'E18',    # Adjust cell reference as needed
                'lam3_processTime': 'E19',    # Adjust cell reference as needed
            }
            
            # Fill lamination parameters
            for lam_key, lam_data in lam_params.items():
                if lam_key == 'lam1':
                    if 'pumpingTime' in lam_data:
                        worksheet[lam_params_mapping['lam1_pumpingTime']] = lam_data['pumpingTime']
                    if 'pressingTime' in lam_data:
                        worksheet[lam_params_mapping['lam1_pressingTime']] = lam_data['pressingTime']
                    if 'ventingTime' in lam_data:
                        worksheet[lam_params_mapping['lam1_ventingTime']] = lam_data['ventingTime']
                    if 'processTime' in lam_data:
                        worksheet[lam_params_mapping['lam1_processTime']] = lam_data['processTime']
                        
                elif lam_key == 'lam2':
                    if 'pumpingTime' in lam_data:
                        worksheet[lam_params_mapping['lam2_pumpingTime']] = lam_data['pumpingTime']
                    if 'pressingTime' in lam_data:
                        worksheet[lam_params_mapping['lam2_pressingTime']] = lam_data['pressingTime']
                    if 'ventingTime' in lam_data:
                        worksheet[lam_params_mapping['lam2_ventingTime']] = lam_data['ventingTime']
                    if 'processTime' in lam_data:
                        worksheet[lam_params_mapping['lam2_processTime']] = lam_data['processTime']
                        
                elif lam_key == 'lam3':
                    if 'pumpingTime' in lam_data:
                        worksheet[lam_params_mapping['lam3_pumpingTime']] = lam_data['pumpingTime']
                    if 'pressingTime' in lam_data:
                        worksheet[lam_params_mapping['lam3_pressingTime']] = lam_data['pressingTime']
                    if 'ventingTime' in lam_data:
                        worksheet[lam_params_mapping['lam3_ventingTime']] = lam_data['ventingTime']
                    if 'processTime' in lam_data:
                        worksheet[lam_params_mapping['lam3_processTime']] = lam_data['processTime']
            
            log_progress(f"Lamination parameters filled successfully")
            
        log_progress("Basic adhesion test information filled successfully")
    except Exception as e:
        log_progress(f"Error filling basic adhesion test info: {str(e)}")
        raise

def fill_adhesion_test_data(worksheet, adhesion_data):
    try:
        form_data = adhesion_data.get('form_data', {})
        averages = {
            **adhesion_data.get('averages', {}),
            **build_adhesion_averages(form_data)
        }
        
        test_data_mapping = {
            'adhesion_data_0': 'B31',  # Position 1 Front Min
            'adhesion_data_1': 'C31',  # Position 1 Front Max
            'adhesion_data_2': 'D31',  # Position 1 Back Min
            'adhesion_data_3': 'E31',  # Position 1 Back Max
            'adhesion_data_4': 'B32',  # Position 2 Front Min
            'adhesion_data_5': 'C32',  # Position 2 Front Max
            'adhesion_data_6': 'D32',  # Position 2 Back Min
            'adhesion_data_7': 'E32',  # Position 2 Back Max
            'adhesion_data_8': 'B33',  # Position 3 Front Min
            'adhesion_data_9': 'C33',  # Position 3 Front Max
            'adhesion_data_10': 'D33', # Position 3 Back Min
            'adhesion_data_11': 'E33', # Position 3 Back Max
            'adhesion_data_12': 'B34', # Position 4 Front Min
            'adhesion_data_13': 'C34', # Position 4 Front Max
            'adhesion_data_14': 'D34', # Position 4 Back Min
            'adhesion_data_15': 'E34', # Position 4 Back Max
            'adhesion_data_16': 'B35', # Position 5 Front Min
            'adhesion_data_17': 'C35', # Position 5 Front Max
            'adhesion_data_18': 'D35', # Position 5 Back Min
            'adhesion_data_19': 'E35', # Position 5 Back Max
        }
        
        for field, cell_ref in test_data_mapping.items():
            if field in form_data:
                value = form_data[field]
                cell = worksheet[cell_ref]

                if value == '-' or value == '':
                    cell.value = '-'
                else:
                    cell.value = value
                    numeric_value = parse_adhesion_numeric_value(value)
                    threshold = get_adhesion_threshold(field)
                    if numeric_value is not None and threshold is not None and numeric_value < threshold:
                        apply_adhesion_fail_style(cell)
        
        averages_mapping = {
            'frontMinAvg': 'B36',
            'frontMaxAvg': 'C36',
            'backMinAvg': 'D36',
            'backMaxAvg': 'E36'
        }
        
        for field, cell_ref in averages_mapping.items():
            if field in averages:
                value = averages[field]
                worksheet[cell_ref] = value if value not in (None, '', '-') else '0.00'

        log_progress("Adhesion test data filled successfully")
    except Exception as e:
        log_progress(f"Error filling adhesion test data: {str(e)}")
        raise

def generate_adhesion_filename(adhesion_data):
    report_name = adhesion_data.get('name', 'Adhesion_Test_Report')
    timestamp = adhesion_data.get('timestamp', '').split('T')[0] if adhesion_data.get('timestamp') else ''
    clean_name = "".join(c for c in report_name if c.isalnum() or c in (' ', '-', '_')).rstrip()
    clean_name = clean_name.replace(' ', '_')
    if timestamp:
        return f"Adhesion_Test_{clean_name}_{timestamp}.xlsx"
    else:
        return f"Adhesion_Test_{clean_name}.xlsx"

def generate_adhesion_report(adhesion_data):
    try:
        if not adhesion_data:
            raise ValueError("No adhesion test data provided")
        log_progress("Received adhesion test data for report generation")
        log_progress(f"Report name: {adhesion_data.get('name', 'N/A')}")
        log_progress(f"Form data keys: {list(adhesion_data.get('form_data', {}).keys())}")
        log_progress(f"Averages keys: {list(adhesion_data.get('averages', {}).keys())}")
        
        template_key = get_template_key('Blank Adhesion Test Report.xlsx')
        template_path = download_from_s3(template_key)
        wb = load_workbook(template_path)
        ws = wb.active
        
        fill_adhesion_basic_info(ws, adhesion_data)
        fill_adhesion_test_data(ws, adhesion_data)
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = generate_adhesion_filename(adhesion_data)
        log_progress(f"Adhesion test report generated successfully: {filename}")
        return output, filename
    except Exception as e:
        log_progress(f"Error generating adhesion test report: {str(e)}")
        raise

