from openpyxl import load_workbook
from openpyxl.styles import (PatternFill, Alignment, NamedStyle)
from openpyxl.utils import get_column_letter
import base64
import copy
import io
import os
import urllib.request
import re
from urllib.parse import unquote, urlparse
from openpyxl.cell.cell import MergedCell
from openpyxl.drawing.image import Image as OpenpyxlImage
from generators.audit_mappings import get_audit_field_config, get_audit_observation_mapping, get_audit_template_filename
from generators.audit_mappings.mapping_helpers import (
    AUTO_BUSSING_PATCH_PARAMETER_ID,
    LINE_I_GROUPED_LINE_SELECTION_CELL_MAPPINGS,
    LINE_I_LINE_SELECTION_CELL_MAPPINGS,
)
from paths import get_template_key, download_from_s3
from models.ipqc_audit_models import normalize_ipqc_audit_data
from s3_service import S3Service

DYNAMIC_LINE_PARAMETER_IDS = {
    '2-4', '2-5', '2-6',
    '3-7', '3-8', '3-9',
    '9-6', '9-7', '9-8',
    '10-4', '10-5', '10-6',
    '11-3', '11-4', '11-5',
    '12-1', '12-2', '16-1', '23-1', '27-1', '29-1',
}

GROUPED_SAMPLE_PARAMETER_IDS = {'12-1', '12-2', '16-1', '23-1', '27-1', '29-1'}
SAMPLE_GROUP_ORDER = ('2h', '4h', '6h', '8h')
SAMPLE_GROUPS = {
    '2h': [('Sample-1', 'Sample-1'), ('Sample-2', 'Sample-2'), ('Sample-3', 'Sample-3'), ('Sample-4', 'Sample-4'), ('Sample-5', 'Sample-5')],
    '4h': [('Sample-6', 'Sample-6'), ('Sample-7', 'Sample-7'), ('Sample-8', 'Sample-8'), ('Sample-9', 'Sample-9'), ('Sample-10', 'Sample-10')],
    '6h': [('Sample-11', 'Sample-1'), ('Sample-12', 'Sample-2'), ('Sample-13', 'Sample-3'), ('Sample-14', 'Sample-4'), ('Sample-15', 'Sample-5')],
    '8h': [('Sample-16', 'Sample-6'), ('Sample-17', 'Sample-7'), ('Sample-18', 'Sample-8'), ('Sample-19', 'Sample-9'), ('Sample-20', 'Sample-10')],
}
GROUPED_SAMPLE_TARGET_SLOTS = {
    '2h': {'lineIndex': 0, 'sampleKeys': ('Sample-1', 'Sample-2', 'Sample-3', 'Sample-4', 'Sample-5')},
    '4h': {'lineIndex': 0, 'sampleKeys': ('Sample-6', 'Sample-7', 'Sample-8', 'Sample-9', 'Sample-10')},
    '6h': {'lineIndex': 1, 'sampleKeys': ('Sample-1', 'Sample-2', 'Sample-3', 'Sample-4', 'Sample-5')},
    '8h': {'lineIndex': 1, 'sampleKeys': ('Sample-6', 'Sample-7', 'Sample-8', 'Sample-9', 'Sample-10')},
}
LINE_SELECTION_CELL_MAPPINGS = {
    '2-4': {'4h': 'M16', '8h': 'Y16'}, '2-5': {'4h': 'M16', '8h': 'Y16'}, '2-6': {'4h': 'M16', '8h': 'Y16'},
    '3-7': {'4h': 'M27', '8h': 'Y27'}, '3-8': {'4h': 'M27', '8h': 'Y27'}, '3-9': {'4h': 'M27', '8h': 'Y27'},
    '9-6': {'4h': 'M191', '8h': 'Y191'}, '9-7': {'4h': 'M191', '8h': 'Y191'}, '9-8': {'4h': 'M191', '8h': 'Y191'},
    '10-4': {'4h': 'M199', '8h': 'Y199'}, '10-5': {'4h': 'M199', '8h': 'Y199'}, '10-6': {'4h': 'M199', '8h': 'Y199'},
    '11-3': {'4h': 'M203', '8h': 'Y203'}, '11-4': {'4h': 'M203', '8h': 'Y203'}, '11-5': {'4h': 'M203', '8h': 'Y203'},
}
GROUPED_LINE_SELECTION_CELL_MAPPINGS = {
    '12-1': {'2h': 'J211', '4h': 'P211', '6h': 'V211', '8h': 'AB211'},
    '12-2': {'2h': 'J211', '4h': 'P211', '6h': 'V211', '8h': 'AB211'},
    '16-1': {'2h': 'J250', '4h': 'P250', '6h': 'V250', '8h': 'AB250'},
    '23-1': {'2h': 'J332', '4h': 'P332', '6h': 'V332', '8h': 'AB332'},
    '27-1': {'2h': 'J375', '4h': 'P375', '6h': 'V375', '8h': 'AB375'}
}
SIMPLE_DEFAULTS = {
    '1-3': 'Checked OK', '2-3': 'Checked OK', '3-4': 'New Box Open',
    '4-2': 'New Box Open', '4-3': 'Stacked two box of cells', '4-4': 'Checked OK',
    '5-2': 'Checked OK',
    '6-1': 'Checked OK', '13-1': 'Checked OK', '13-2': 'Checked OK',
    '13-4': 'Checked OK', '25-1': 'Checked OK', '25-2': 'Checked OK',
    '25-3': 'Checked OK', '28-1': 'Checked OK', '29-2': 'Checked OK', '30-1': 'Checked OK',
    '30-2': 'Checked OK',
}
SAMPLE_KEY_DEFAULTS = {
    '2-2': 'Checked OK', '3-6': 'Checked OK', '9-5': 'Checked OK', '11-2': 'Checked OK',
}
LINE_KEY_DEFAULTS = {
    '7-5': 'Checked OK', '8-1': 'Checked OK', '8-2': 'Outside RFID',
    '8-3': 'Checked OK', '8-4': 'Checked OK', '8-5': 'Checked OK',
    '15-1': 'Checked OK',
    '15-2': 'Checked OK', '20-2': 'Checked OK', '20-4': 'Checked OK',
    '21-5': 'Checked OK', '22-1': 'Checked OK', '22-2': 'Checked OK',
    '17-3': 'Both side of length', '17-4': 'Checked OK',
    '18-4': 'Checked OK', '19-1': 'ESD band used',
    '24-2': 'Checked OK', '24-3': 'Checked OK', '24-8': 'Checked OK',
    '24-9': 'Checked OK', '26-1': 'Pass', '26-2': 'Pass', '26-3': 'Pass',
}
SAFETY_TEST_PARAM_IDS = {'26-1', '26-2', '26-3'}
SAFETY_TIME_KEYS = ('4hrs', '8hrs')
SAFETY_MODULE_FIELDS = {
    'lineA_4hr_moduleId',
    'lineA_8hr_moduleId',
    'lineB_4hr_moduleId',
    'lineB_8hr_moduleId',
}
LEGACY_SAFETY_MODULE_FIELDS_BY_TIME = {
    '4hrs': 'moduleId4Hours',
    '8hrs': 'moduleId8Hours',
}
SAFETY_MODULE_FIELD_BY_LINE_TIME = {
    ('Line-1', '4hrs'): 'lineA_4hr_moduleId',
    ('Line-1', '8hrs'): 'lineA_8hr_moduleId',
    ('Line-3', '4hrs'): 'lineA_4hr_moduleId',
    ('Line-3', '8hrs'): 'lineA_8hr_moduleId',
    ('Line-2', '4hrs'): 'lineB_4hr_moduleId',
    ('Line-2', '8hrs'): 'lineB_8hr_moduleId',
    ('Line-4', '4hrs'): 'lineB_4hr_moduleId',
    ('Line-4', '8hrs'): 'lineB_8hr_moduleId',
}
SAFETY_LINE_PAIRS = (('Line-1', 'Line-3'), ('Line-2', 'Line-4'))
SAFETY_TEMPLATE_LINE_ALIAS = {
    'Line-1': 'Line-3',
    'Line-2': 'Line-4',
}
TIME_SLOT_DEFAULTS = {
    '10-1': 'OFF', '10-2': 'OFF', '10-3': 'OFF',
    '31-2': 'Checked OK', '31-3': 'Checked OK', '31-4': 'Checked OK',
    '31-5': 'Checked OK', '31-6': 'Checked OK', '31-7': 'Checked OK',
    '31-8': 'Vertically', '31-10': 'Checked OK', '31-11': 'Checked OK',
    '31-12': 'Checked OK', '31-13': 'Checked OK',
}
SAMPLE_OFF_DEFAULTS = {'10-4', '10-5', '10-6'}
COMPONENT_FIELD_DEFAULTS = {
    '18-1': {
        'type': 'TM4545-30U/TM3045-30U/XND-18-V100C/TM45100-30U'
    }
}
NESTED_STRING_DEFAULTS = {
    '5-5-cell-appearance': 'Checked OK',
    '5-15-el-inspection': 'Checked OK',
}
NUMERIC_PATTERN = re.compile(r'^[+-]?(?:\d+(?:\.\d+)?|\.\d+)$')
SIGNATURE_S3_PREFIX = 'users/signatures/'
AUTO_TAPING_LEGACY_LINE_BY_INTERNAL = {
    'line1_auto_taping_1': 'Line-1',
    'line1_auto_taping_2': 'Line-2',
    'line1_auto_taping_3': 'Line-3',
    'line2_auto_taping_3': 'Line-3',
    'line2_auto_taping_4': 'Line-4',
}
AUTO_TAPING_INTERNAL_LINES_BY_AUDIT_LINE = {
    'I': ('line1_auto_taping_1', 'line1_auto_taping_2', 'line1_auto_taping_3'),
    'II': ('line2_auto_taping_3', 'line2_auto_taping_4'),
}
AUTO_TAPING_EXPORT_FIELDS = ('4hrs', '8hrs', 'Supplier', 'Type', 'Quantity')

def is_empty_value(value):
    return value is None or value == ''

def is_safety_module_id_field(key):
    return key in SAFETY_MODULE_FIELDS or key in LEGACY_SAFETY_MODULE_FIELDS_BY_TIME.values()

def split_safety_result_key(key):
    if not isinstance(key, str):
        return None, None
    line_key, separator, time_key = key.rpartition('-')
    if not separator or time_key not in SAFETY_TIME_KEYS:
        return None, None
    return line_key, time_key

def apply_safety_export_defaults(param_id, value, line_number):
    normalized_value = {} if not isinstance(value, dict) else dict(value)
    default_value = LINE_KEY_DEFAULTS[param_id]
    target_line_index = 0 if line_number == 'I' else 1

    for line_pair in SAFETY_LINE_PAIRS:
        target_line = line_pair[target_line_index]
        for time_key in SAFETY_TIME_KEYS:
            equivalent_keys = [f"{line}-{time_key}" for line in line_pair]
            if any(not is_empty_value(normalized_value.get(key)) for key in equivalent_keys):
                continue
            normalized_value[f"{target_line}-{time_key}"] = default_value
    return normalized_value

def normalize_auto_taping_export_value(value, line_number=None):
    if not isinstance(value, dict):
        return value

    normalized_value = dict(value)
    internal_lines = AUTO_TAPING_INTERNAL_LINES_BY_AUDIT_LINE.get(
        line_number,
        tuple(AUTO_TAPING_LEGACY_LINE_BY_INTERNAL),
    )
    for internal_line in internal_lines:
        legacy_line = AUTO_TAPING_LEGACY_LINE_BY_INTERNAL[internal_line]
        for field_name in AUTO_TAPING_EXPORT_FIELDS:
            internal_key = f"{internal_line}-{field_name}"
            legacy_key = f"{legacy_line}-{field_name}"
            if internal_key in normalized_value and is_empty_value(normalized_value.get(legacy_key)):
                normalized_value[legacy_key] = normalized_value.get(internal_key, '')
    return normalized_value

def apply_export_defaults(param_id, value, param_mapping, line_number=None):
    """Match frontend selector defaults during export without overwriting user input."""
    if param_id in SIMPLE_DEFAULTS and is_empty_value(value):
        return SIMPLE_DEFAULTS[param_id]
    if param_id in SAMPLE_KEY_DEFAULTS:
        normalized_value = {} if not isinstance(value, dict) else dict(value)
        for sample_key in ('Sample-1', 'Sample-2', 'Sample-3', 'Sample-4', 'Sample-5', 'Sample-6'):
            if is_empty_value(normalized_value.get(sample_key)):
                normalized_value[sample_key] = SAMPLE_KEY_DEFAULTS[param_id]
        return normalized_value
    if param_id in SAMPLE_OFF_DEFAULTS:
        normalized_value = {} if not isinstance(value, dict) else dict(value)
        for line_mapping in param_mapping.values():
            if not isinstance(line_mapping, dict):
                continue
            for sample_key in line_mapping:
                if is_empty_value(normalized_value.get(sample_key)):
                    normalized_value[sample_key] = 'OFF'
        return normalized_value
    if param_id in TIME_SLOT_DEFAULTS and is_empty_value(value):
        return TIME_SLOT_DEFAULTS[param_id]
    if param_id in COMPONENT_FIELD_DEFAULTS and isinstance(param_mapping, dict):
        normalized_value = {} if not isinstance(value, dict) else dict(value)
        for key in param_mapping:
            for field_name, default_value in COMPONENT_FIELD_DEFAULTS[param_id].items():
                if key.endswith(f"-{field_name}") and is_empty_value(normalized_value.get(key)):
                    normalized_value[key] = default_value
        return normalized_value
    if param_id in NESTED_STRING_DEFAULTS and isinstance(param_mapping, dict):
        return apply_nested_string_default(value, param_mapping, NESTED_STRING_DEFAULTS[param_id])
    if param_id in SAFETY_TEST_PARAM_IDS:
        return apply_safety_export_defaults(param_id, value, line_number)
    if param_id not in LINE_KEY_DEFAULTS or not isinstance(param_mapping, dict):
        return value

    normalized_value = {} if not isinstance(value, dict) else dict(value)
    default_value = LINE_KEY_DEFAULTS[param_id]
    for key in param_mapping:
        if is_safety_module_id_field(key):
            continue
        if is_empty_value(normalized_value.get(key)):
            normalized_value[key] = default_value
    return normalized_value

def apply_nested_string_default(value, param_mapping, default_value):
    """Hydrate known nested selector defaults while preserving explicit values."""
    normalized_value = {} if not isinstance(value, dict) else dict(value)
    for outer_key, inner_mapping in param_mapping.items():
        if not isinstance(inner_mapping, dict):
            continue
        current_inner = normalized_value.get(outer_key)
        current_inner = dict(current_inner) if isinstance(current_inner, dict) else {}
        for inner_key in inner_mapping:
            if is_empty_value(current_inner.get(inner_key)):
                current_inner[inner_key] = default_value
        normalized_value[outer_key] = current_inner
    return normalized_value

def combine_result_with_module_id(result, module_id):
    if module_id:
        return f"{result} - {module_id}" if result else module_id
    return result

def normalize_safety_module_id_source(value):
    if not isinstance(value, dict):
        return {}

    source = dict(value)
    for field in SAFETY_MODULE_FIELDS:
        if field in source and source.get(field) is not None:
            continue
        legacy_key = LEGACY_SAFETY_MODULE_FIELDS_BY_TIME['4hrs' if '_4hr_' in field else '8hrs']
        source[field] = source.get(legacy_key, '') or ''
    return source

def get_safety_module_id_for_result(module_source, result_key):
    if not isinstance(module_source, dict):
        return ''

    line_key, time_key = split_safety_result_key(result_key)
    module_field = SAFETY_MODULE_FIELD_BY_LINE_TIME.get((line_key, time_key))
    if module_field:
        if module_field in module_source:
            return module_source.get(module_field, '') or ''

    legacy_key = LEGACY_SAFETY_MODULE_FIELDS_BY_TIME.get(time_key)
    return (module_source.get(legacy_key, '') or '') if legacy_key else ''

def get_safety_mapping_key(key, param_mapping):
    if key in param_mapping:
        return key

    line_key, time_key = split_safety_result_key(key)
    alias_line = SAFETY_TEMPLATE_LINE_ALIAS.get(line_key)
    if not alias_line:
        return None

    alias_key = f"{alias_line}-{time_key}"
    return alias_key if alias_key in param_mapping else None

def parse_numeric(value):
    """Write numeric-only strings to Excel as real numeric cells."""
    if isinstance(value, str):
        stripped = value.strip()
        if not stripped or not NUMERIC_PATTERN.fullmatch(stripped):
            return value
        return float(stripped) if '.' in stripped else int(stripped)
    return value

def set_cell_value(cell, value):
    cell.value = parse_numeric(value)

def set_cell_raw_value(cell, value):
    cell.value = value

def is_standard_sample_grouped_value(value):
    return isinstance(value, dict) and isinstance(value.get('sampleGroups'), list)

def sample_number_from_label(sample_label):
    if not isinstance(sample_label, str) or not sample_label.startswith('Sample-'):
        return None
    try:
        return int(sample_label.split('-', 1)[1])
    except (TypeError, ValueError):
        return None

def get_grouped_sample_groups(value):
    if not is_standard_sample_grouped_value(value):
        return {}
    groups = {}
    for group in value.get('sampleGroups', []):
        if isinstance(group, dict) and group.get('groupKey') in SAMPLE_GROUPS:
            groups[group.get('groupKey')] = group
    return groups

def get_grouped_sample_line_mapping(value):
    groups = get_grouped_sample_groups(value)
    return {
        group_key: normalize_line_value(groups[group_key].get('selectedLine', ''))
        for group_key in SAMPLE_GROUP_ORDER
        if group_key in groups and groups[group_key].get('selectedLine')
    }

def get_line_sort_key(line_key):
    normalized_line = normalize_line_value(line_key)
    return (0, int(normalized_line)) if str(normalized_line).isdigit() else (1, str(normalized_line))

def get_grouped_sample_mapping_lines(param_mapping):
    if not isinstance(param_mapping, dict):
        return []
    return sorted(
        (line_key for line_key, sample_mapping in param_mapping.items() if isinstance(sample_mapping, dict) and str(line_key).startswith('Line-')),
        key=get_line_sort_key
    )

def get_grouped_sample_cell_map(param_mapping):
    mapping_lines = get_grouped_sample_mapping_lines(param_mapping)
    if len(mapping_lines) < 2:
        print("Warning: grouped sample mapping expected two line blocks but found fewer")
        return {}

    cell_map = {}
    for group_key in SAMPLE_GROUP_ORDER:
        target_slot = GROUPED_SAMPLE_TARGET_SLOTS[group_key]
        target_line = mapping_lines[target_slot['lineIndex']]
        target_line_mapping = param_mapping.get(target_line, {})
        for (source_sample, _), target_sample in zip(SAMPLE_GROUPS[group_key], target_slot['sampleKeys']):
            sample_number = sample_number_from_label(source_sample)
            cell_ref = target_line_mapping.get(target_sample) if isinstance(target_line_mapping, dict) else None
            if isinstance(sample_number, int) and cell_ref:
                cell_map[sample_number] = cell_ref
    return cell_map

def get_grouped_sample_lookup(group):
    lookup = {}
    duplicate_indexes = set()
    samples = group.get('samples', []) if isinstance(group, dict) else []
    for sample in samples:
        if not isinstance(sample, dict):
            continue
        sample_number = sample.get('sampleNumber')
        if not isinstance(sample_number, int):
            sample_number = sample_number_from_label(sample.get('sampleLabel'))
        if isinstance(sample_number, int) and sample_number not in lookup:
            lookup[sample_number] = sample
        elif isinstance(sample_number, int):
            duplicate_indexes.add(sample_number)
    if duplicate_indexes:
        group_key = group.get('groupKey', '') if isinstance(group, dict) else ''
        print(f"Warning: grouped sample group {group_key} has duplicate sample indexes: {', '.join(str(index) for index in sorted(duplicate_indexes))}")
    return lookup

def validate_grouped_sample_value(param_id, value):
    if not is_standard_sample_grouped_value(value):
        print(f"Warning: grouped sample parameter {param_id} has malformed value; expected sampleGroups structure")
        return

    groups = get_grouped_sample_groups(value)
    missing_groups = [group_key for group_key in SAMPLE_GROUP_ORDER if group_key not in groups]
    if missing_groups:
        print(f"Warning: grouped sample parameter {param_id} missing groups: {', '.join(missing_groups)}")

    total_samples = 0
    for group_key in SAMPLE_GROUP_ORDER:
        group = groups.get(group_key)
        if not group:
            continue
        lookup = get_grouped_sample_lookup(group)
        expected_numbers = [sample_number_from_label(source_sample) for source_sample, _ in SAMPLE_GROUPS[group_key]]
        missing_samples = [f"Sample-{sample_number}" for sample_number in expected_numbers if sample_number not in lookup]
        if missing_samples:
            print(f"Warning: grouped sample parameter {param_id} group {group_key} missing samples: {', '.join(missing_samples)}")
        if not group.get('selectedLine'):
            print(f"Warning: grouped sample parameter {param_id} group {group_key} missing selectedLine")
        actual_numbers = sorted(lookup)
        if actual_numbers != [number for number in expected_numbers if isinstance(number, int)]:
            print(f"Warning: grouped sample parameter {param_id} group {group_key} has unexpected sample order/indexes: {actual_numbers}")
        total_samples += len(lookup)

    if total_samples != 20:
        print(f"Warning: grouped sample parameter {param_id} expected 20 samples, found {total_samples}")

def extract_safety_module_ids(stages):
    for stage in stages:
        if stage.get('id') != 26:
            continue
        for parameter in stage.get('parameters', []):
            if parameter.get('id') != '26-1':
                continue
            for observation in parameter.get('observations', []):
                value = observation.get('value', {})
                if isinstance(value, dict):
                    return normalize_safety_module_id_source(value)
    return {}

def get_writable_cell(worksheet, cell_ref):
    try:
        cell = worksheet[cell_ref]
        if isinstance(cell, MergedCell):
            for merged_range in worksheet.merged_cells.ranges:
                if cell.coordinate in merged_range:
                    return worksheet[merged_range.coord.split(":")[0]]
        return cell
    except Exception as e:
        print(f"Warning: invalid or inaccessible Excel cell reference {cell_ref}: {str(e)}")
        return None

def safe_set_cell(worksheet, cell_ref, value, raw=False, **format_options):
    cell = get_writable_cell(worksheet, cell_ref)
    if cell is None:
        return False
    if raw:
        set_cell_raw_value(cell, value)
    else:
        set_cell_value(cell, value)
    apply_cell_formatting(cell, **format_options)
    return True

def extract_signature_image_key(image_source):
    if not isinstance(image_source, str):
        return None

    source = image_source.strip()
    if source.startswith(SIGNATURE_S3_PREFIX):
        return source

    if source.startswith(('http://', 'https://')):
        parsed = urlparse(source)
        path_key = unquote(parsed.path.lstrip('/'))
        marker_index = path_key.find(SIGNATURE_S3_PREFIX)
        if marker_index >= 0:
            return path_key[marker_index:]

    return None

def describe_signature_image_source(image_source):
    if not image_source:
        return 'missing'
    if not isinstance(image_source, str):
        return type(image_source).__name__
    if image_source.startswith('data:'):
        return 'data URI'
    if image_source.startswith(('http://', 'https://')):
        signature_key = extract_signature_image_key(image_source)
        return f"S3 presigned URL ({signature_key})" if signature_key else 'URL'
    if image_source.startswith(SIGNATURE_S3_PREFIX):
        return f"S3 key ({image_source})"
    if os.path.exists(image_source):
        return f"file path ({image_source})"
    return 'base64/reference'

def load_signature_image_from_s3(signature_key):
    try:
        s3_service = S3Service()
        response = s3_service.s3_client.get_object(
            Bucket=s3_service.bucket_name,
            Key=signature_key
        )
        return io.BytesIO(response['Body'].read())
    except Exception as e:
        print(f"Warning: signature image S3 reference is not accessible ({signature_key}): {str(e)}")
        return None

def resolve_signature_image_bytes(image_source):
    if not image_source:
        return None

    signature_key = extract_signature_image_key(image_source)
    if signature_key:
        return load_signature_image_from_s3(signature_key)

    if isinstance(image_source, str) and image_source.startswith(('http://', 'https://')):
        with urllib.request.urlopen(image_source, timeout=10) as response:
            return io.BytesIO(response.read())

    if isinstance(image_source, str) and os.path.exists(image_source):
        return image_source

    encoded_image = image_source.split(',', 1)[1] if isinstance(image_source, str) and ',' in image_source else image_source
    return io.BytesIO(base64.b64decode(encoded_image))

def insert_signature_image(worksheet, cell_ref, image_source):
    """Insert URL, data URI, raw base64, or file-path signature images."""
    if not image_source:
        return False
    try:
        image_bytes = resolve_signature_image_bytes(image_source)
        if not image_bytes:
            print(f"Warning: signature image at {cell_ref} could not be loaded from {describe_signature_image_source(image_source)}")
            return False
        signature_image = OpenpyxlImage(image_bytes)
        signature_image.height = 45
        signature_image.width = 120
        worksheet.add_image(signature_image, cell_ref)
        return True
    except Exception as e:
        print(f"Unable to insert signature image at {cell_ref}: {str(e)}")
        return False

def setup_cell_styles(workbook):
    header_style = NamedStyle(name="header_style")
    header_style.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_style.alignment = Alignment(horizontal='center', vertical='center')
    data_style = NamedStyle(name="data_style")
    data_style.alignment = Alignment(horizontal='center', vertical='center')
    important_style = NamedStyle(name="important_style")
    important_style.alignment = Alignment(horizontal='center', vertical='center')
    for style in [header_style, data_style, important_style]:
        if style.name not in workbook.named_styles:
            workbook.add_named_style(style)

def apply_cell_formatting(cell, horizontal='center', vertical='center', **_format_options):
    cell.alignment = Alignment(
        horizontal=horizontal,
        vertical=vertical,
        wrap_text=True
    )


def get_template_config(line_number):
    if line_number not in ('I', 'II'):
        raise ValueError(f"Unsupported line number: {line_number}")

    template_key = get_template_key(get_audit_template_filename(line_number))
    template_path = download_from_s3(template_key)
    field_config = get_audit_field_config(line_number)
    observation_cell_mapping = get_audit_observation_mapping(line_number)
    return template_path, field_config, observation_cell_mapping

def resolve_signature(audit_data, signature_key):
    """Read current and legacy signature shapes without requiring schema changes."""
    signatures = audit_data.get('signatures') or audit_data.get('data', {}).get('signatures') or {}
    image_key = f"{signature_key}Image"
    nested_data = audit_data.get('data', {}) if isinstance(audit_data.get('data'), dict) else {}
    signature_value = signatures.get(signature_key) or audit_data.get(signature_key) or nested_data.get(signature_key)
    image_value = signatures.get(image_key) or audit_data.get(image_key) or nested_data.get(image_key)

    if isinstance(signature_value, dict):
        text_value = signature_value.get('name') or signature_value.get('text') or signature_value.get('value') or ''
        image_value = image_value or signature_value.get('signature') or signature_value.get('image') or signature_value.get('photo')
    else:
        text_value = signature_value or ''

    return text_value, image_value

def write_signature(worksheet, text_cell_ref, image_cell_ref, text_value, image_value, format_options, signature_label='Signature'):
    if text_value:
        text_cell = get_writable_cell(worksheet, text_cell_ref)
        if text_cell is not None:
            text_cell.value = text_value
            apply_cell_formatting(text_cell, **{**format_options, 'horizontal': 'left'})

    if image_value:
        print(f"{signature_label} image reference found: {describe_signature_image_source(image_value)}")
        image_cell = get_writable_cell(worksheet, image_cell_ref)
        if insert_signature_image(worksheet, image_cell_ref, image_value):
            if image_cell is not None:
                image_cell.value = None
                apply_cell_formatting(image_cell, **format_options)
    elif text_value:
        print(f"Warning: {signature_label} text signature exists but image reference is missing; continuing without image")

def normalize_line_value(line_value):
    if line_value is None:
        return ''
    line_text = str(line_value).strip()
    return line_text.replace('Line-', '', 1) if line_text.startswith('Line-') else line_text

def format_line_key(line_value):
    normalized_value = normalize_line_value(line_value)
    return f"Line-{normalized_value}" if normalized_value else ''

def get_line_selection_bucket(time_slot):
    selected_slot = normalize_line_value(time_slot)
    if selected_slot in ('1', '3'):
        return '4h'
    if selected_slot in ('2', '4'):
        return '8h'
    return ''

def fill_line_dropdown_values(worksheet, audit_data):
    """Export selected line dropdowns into their template header cells."""
    try:
        is_line_i = audit_data.get('lineNumber') == 'I'
        line_selection_mappings = LINE_I_LINE_SELECTION_CELL_MAPPINGS if is_line_i else LINE_SELECTION_CELL_MAPPINGS
        grouped_selection_mappings = (
            LINE_I_GROUPED_LINE_SELECTION_CELL_MAPPINGS
            if is_line_i
            else GROUPED_LINE_SELECTION_CELL_MAPPINGS
        )
        for stage in audit_data.get('stages', []) or []:
            if not isinstance(stage, dict):
                continue
            for parameter in stage.get('parameters', []) or []:
                if not isinstance(parameter, dict):
                    continue
                param_id = parameter.get('id', '')
                for observation in parameter.get('observations', []) or []:
                    if not isinstance(observation, dict):
                        continue
                    selected_line = observation.get('selectedLine')
                    line_cells = line_selection_mappings.get(param_id)
                    if selected_line and line_cells:
                        bucket = get_line_selection_bucket(observation.get('timeSlot', ''))
                        cell_ref = line_cells.get(bucket)
                        if cell_ref:
                            safe_set_cell(worksheet, cell_ref, normalize_line_value(selected_line), raw=True, horizontal='center')

                    group_cells = grouped_selection_mappings.get(param_id)
                    if group_cells:
                        grouped_line_mapping = get_grouped_sample_line_mapping(observation.get('value', {}))
                        for group_key in SAMPLE_GROUP_ORDER:
                            line_value = grouped_line_mapping.get(group_key)
                            cell_ref = group_cells.get(group_key)
                            if line_value and cell_ref:
                                safe_set_cell(worksheet, cell_ref, normalize_line_value(line_value), raw=True, horizontal='center')
    except Exception as e:
        print(f"Warning: failed to export some line dropdown values: {str(e)}")

def fill_basic_info(worksheet, audit_data, field_config):
    try:
        if audit_data.get('date'):
            cell_ref = field_config['date']['cell']
            safe_set_cell(worksheet, cell_ref, audit_data['date'], raw=True, **field_config['date'].get('format', {}))

        if audit_data.get('shift'):
            cell_ref = field_config['shift']['cell']
            shift_map = {'A': 'A', 'B': 'B', 'C': 'C', 'G': 'G'}
            safe_set_cell(worksheet, cell_ref, shift_map.get(audit_data['shift'], audit_data['shift']), raw=True, **field_config['shift'].get('format', {}))

        if audit_data.get('productionOrderNo'):
            cell_ref = field_config['production_order']['cell']
            safe_set_cell(worksheet, cell_ref, audit_data['productionOrderNo'], raw=True, **field_config['production_order'].get('format', {}))

        if audit_data.get('moduleType'):
            cell_ref = field_config['module_type']['cell']
            safe_set_cell(worksheet, cell_ref, audit_data['moduleType'], raw=True, **field_config['module_type'].get('format', {}))

        audit_by_text, audit_by_image = resolve_signature(audit_data, 'auditBy')
        audit_by_cell = field_config.get('audit_by', {}).get('cell', 'C405')
        audit_by_image_cell = field_config.get('auditByImage', {}).get('cell', 'C407')
        write_signature(worksheet, audit_by_cell, audit_by_image_cell, audit_by_text, audit_by_image, field_config.get('audit_by', {}).get('format', {}), 'Audit By')

        reviewed_by_text, reviewed_by_image = resolve_signature(audit_data, 'reviewedBy')
        reviewed_by_cell = field_config.get('reviewed_by', {}).get('cell', 'O405')
        reviewed_by_image_cell = field_config.get('reviewedByImage', {}).get('cell', 'O407')
        write_signature(worksheet, reviewed_by_cell, reviewed_by_image_cell, reviewed_by_text, reviewed_by_image, field_config.get('reviewed_by', {}).get('format', {}), 'Reviewed By')

        customer_spec = 'Yes' if audit_data.get('customerSpecAvailable') else 'No'
        spec_signed = 'Yes' if audit_data.get('specificationSignedOff') else 'No'

        cell_ref = field_config['customer_spec_available']['cell']
        safe_set_cell(worksheet, cell_ref, customer_spec, raw=True, **field_config['customer_spec_available'].get('format', {}))

        cell_ref = field_config['spec_signed_off']['cell']
        safe_set_cell(worksheet, cell_ref, spec_signed, raw=True, **field_config['spec_signed_off'].get('format', {}))

        print("Basic information filled and formatted successfully")
        
    except Exception as e:
        print(f"Error filling basic info: {str(e)}")
        print("Warning: continuing audit export without some basic info fields")

def fill_observations_data(worksheet, audit_data, observation_cell_mapping):
    try:
        stages = audit_data.get('stages', [])
        if not stages:
            print("No stages data found for observations")
            return
        safety_module_ids = extract_safety_module_ids(stages)
        
        for stage in stages:
            parameters = stage.get('parameters', [])
            for parameter in parameters:
                param_id = parameter.get('id')
                observations = parameter.get('observations', [])
                
                if param_id in observation_cell_mapping:
                    param_mapping = observation_cell_mapping[param_id]
                    
                    for observation in observations:
                        time_slot = observation.get('timeSlot', '')
                        selected_line = observation.get('selectedLine')
                        line_time_slot = format_line_key(selected_line) if selected_line else time_slot
                        if isinstance(param_id, str) and param_id.startswith('8-'):
                            raw_value = normalize_auto_taping_export_value(observation.get('value', ''), audit_data.get('lineNumber'))
                        else:
                            raw_value = observation.get('value', '')
                        value = apply_export_defaults(param_id, raw_value, param_mapping, audit_data.get('lineNumber'))
                        if isinstance(param_id, str) and param_id.startswith('8-'):
                            value = normalize_auto_taping_export_value(value, audit_data.get('lineNumber'))
                        if param_id in ['2-2', '3-6', '9-5', '11-2']:
                            # Sample-based parameters (Sample-1, Sample-2, etc.)
                            handle_sample_based_parameter(worksheet, param_id, time_slot, value, param_mapping)
                        
                        elif param_id in ['2-4', '2-5', '2-6', '9-6', '9-7', '9-8', '10-4', '10-5', '10-6', 
                                         '11-3', '11-4', '11-5', '12-1', '12-2', '16-1', '23-1', '27-1', '29-1']:
                            # Line-based sample parameters
                            mapped_time_slot = line_time_slot if param_id in DYNAMIC_LINE_PARAMETER_IDS and line_time_slot in param_mapping else time_slot
                            handle_line_sample_parameter(worksheet, param_id, mapped_time_slot, value, param_mapping, observation.get('lineMapping'))
                        
                        elif param_id in ['5-4-laser-power', '5-5-cell-appearance']:
                            # Stringer unit parameters
                            handle_stringer_unit_parameter(worksheet, param_id, value, param_mapping)
                        
                        elif param_id == '5-6-cell-width':
                            # Cell width measurements
                            handle_cell_width_parameter(worksheet, value, param_mapping)
                        
                        elif param_id == '5-7-groove-length':
                            # Groove length measurements
                            handle_groove_length_parameter(worksheet, value, param_mapping)
                        
                        elif param_id == '5-9-machine-temp-setup':
                            # Machine temperature setup
                            handle_machine_temp_parameter(worksheet, value, param_mapping)
                        
                        elif param_id == '5-10-light-intensity-time':
                            # Light intensity and time
                            handle_light_intensity_parameter(worksheet, value, param_mapping)
                        
                        elif param_id == '5-11-peel-strength':
                            # Peel strength parameter
                            handle_peel_strength_parameter(worksheet, value, param_mapping)
                        
                        elif param_id in ['5-12-ribbon-flatten', '5-13-string-length', '5-14-cell-to-cell-gap', '5-15-el-inspection',
                                          '15-2', '19-3', '22-2', '24-1', '24-2', '24-5', '24-7', '24-8']:
                            # Stringer-based parameters
                            handle_stringer_based_parameter(worksheet, param_id, value, param_mapping)
                        
                        elif param_id == '7-1':
                            # BUS ribbon status
                            handle_bus_ribbon_status(worksheet, value, param_mapping)
                        
                        elif param_id == '7-2':
                            # Soldering time
                            handle_soldering_time(worksheet, value, param_mapping)
                        
                        elif param_id == AUTO_BUSSING_PATCH_PARAMETER_ID:
                            # Rear encapsulant terminal patch measurements
                            handle_nested_measurements(worksheet, value, param_mapping)

                        elif param_id in ['7-3', '7-4', '7-5', '7-6', '7-7', '7-8']:
                            # Line-based measurements
                            handle_line_measurements(worksheet, param_id, value, param_mapping)
                        
                        elif param_id == '7-9':
                            # Peel strength bus ribbon
                            handle_bus_peel_strength(worksheet, value, param_mapping)
                        
                        elif param_id in ['8-1', '8-2', '8-3', '8-4', '8-5', '8-9', '8-10', '8-11', '8-12', '8-13', 
                                         '8-14', '8-15', '8-16', '15-1', '17-3', '17-6', '17-7', '17-8', '17-9', 
                                         '17-10', '17-11', '17-12-coating-thickness', '19-1', '19-2', '20-5', '26-1',
                                         '26-2', '26-3']:
                            # Time-based line parameters
                            handle_time_based_line_parameter(worksheet, param_id, value, param_mapping, safety_module_ids)
                        
                        elif param_id in ['8-6', '17-1', '17-2', '18-1', '18-2', '18-3', '20-1']:
                            # Component status parameters
                            handle_component_status(worksheet, param_id, value, param_mapping)
                        
                        elif param_id in [
                            '14-1-laminator1', '14-2-laminator2', '14-3-laminator3', '14-4-laminator4',
                            '14-1-laminator5', '14-2-laminator6', '14-3-laminator7', '14-4-laminator8',
                        ]:
                            # Laminator parameters
                            handle_laminator_parameter(worksheet, param_id, value, param_mapping)
                        
                        elif param_id in ['17-4', '18-4']:
                            # Sample-based line parameters
                            handle_sample_line_parameter(worksheet, param_id, value, param_mapping)
                        
                        elif param_id == '17-5':
                            # Frame sealant weight
                            handle_frame_sealant_weight(worksheet, value, param_mapping)
                        
                        elif param_id in ['18-5', '18-6']:
                            # JB measurements
                            handle_jb_measurements(worksheet, param_id, value, param_mapping)
                        
                        elif param_id in ['20-3', '24-6', '24-9', '24-10']:
                            # Complex measurement parameters
                            handle_complex_measurements(worksheet, param_id, value, param_mapping)
                        
                        elif param_id in ['20-2', '20-4', '21-1', '21-2', '21-3', '21-4', '21-5', '22-1', '24-3', '24-4']:
                            # Time-based environmental parameters
                            handle_environmental_parameters(worksheet, param_id, value, param_mapping)
                        
                        elif param_id in ['5-8-tds', '9-9']:
                            # Simple value parameters
                            if time_slot in param_mapping:
                                cell_ref = param_mapping[time_slot]
                                cell = get_writable_cell(worksheet, cell_ref)
                                set_cell_value(cell, value)
                                apply_cell_formatting(cell, horizontal='center')
                        
                        else:
                            # Default handling for simple timeSlot-value pairs
                            if time_slot in param_mapping:
                                cell_ref = param_mapping[time_slot]
                                cell = get_writable_cell(worksheet, cell_ref)
                                set_cell_value(cell, value)
                                apply_cell_formatting(cell, horizontal='center')
                                print(f"Filled observation {value} for parameter {param_id} at {time_slot} in cell {cell_ref}")
        
        print("Observations data filled successfully")
    except Exception as e:
        print(f"Error filling observations data: {str(e)}")
        print("Warning: continuing audit export with partially filled observations")

def handle_sample_based_parameter(worksheet, param_id, time_slot, value, param_mapping):
    """Handle parameters with Sample-1, Sample-2, etc. structure"""
    if isinstance(value, dict):
        for sample_key, sample_value in value.items():
            if sample_key in param_mapping.get(time_slot, {}):
                cell_ref = param_mapping[time_slot][sample_key]
                cell = get_writable_cell(worksheet, cell_ref)
                set_cell_value(cell, sample_value)
                apply_cell_formatting(cell, horizontal='center')
                print(f"Filled observation {sample_value} for parameter {param_id} at {sample_key} in cell {cell_ref}")

def handle_line_sample_parameter(worksheet, param_id, time_slot, value, param_mapping, line_mapping=None):
    """Handle line-based sample parameters"""
    if isinstance(value, dict):
        if param_id in GROUPED_SAMPLE_PARAMETER_IDS:
            validate_grouped_sample_value(param_id, value)
            groups = get_grouped_sample_groups(value)
            sample_cell_map = get_grouped_sample_cell_map(param_mapping)
            for group_key in SAMPLE_GROUP_ORDER:
                group = groups.get(group_key)
                if not group:
                    continue
                sample_lookup = get_grouped_sample_lookup(group)
                for source_sample, _target_sample in SAMPLE_GROUPS[group_key]:
                    sample_number = sample_number_from_label(source_sample)
                    if not isinstance(sample_number, int):
                        continue
                    sample = sample_lookup.get(sample_number)
                    sample_value = sample.get('value', '') if isinstance(sample, dict) else ''
                    cell_ref = sample_cell_map.get(sample_number)
                    if not cell_ref:
                        print(f"Warning: no Excel cell mapping for parameter {param_id} {source_sample}")
                        continue
                    cell = get_writable_cell(worksheet, cell_ref)
                    set_cell_value(cell, sample_value)
                    apply_cell_formatting(cell, horizontal='center')
                    if sample_value != '':
                        print(f"Filled observation {sample_value} for parameter {param_id} at {source_sample} in cell {cell_ref}")
            return

        for sample_key, sample_value in value.items():
            if time_slot in param_mapping and sample_key in param_mapping[time_slot]:
                cell_ref = param_mapping[time_slot][sample_key]
                cell = get_writable_cell(worksheet, cell_ref)
                set_cell_value(cell, sample_value)
                apply_cell_formatting(cell, horizontal='center')
                print(f"Filled observation {sample_value} for parameter {param_id} at {sample_key} in cell {cell_ref}")

def handle_stringer_unit_parameter(worksheet, param_id, value, param_mapping):
    """Handle stringer unit parameters"""
    if isinstance(value, dict):
        for stringer_key, stringer_value in value.items():
            if stringer_key in param_mapping:
                for unit_key, unit_value in stringer_value.items():
                    if unit_key in param_mapping[stringer_key]:
                        cell_ref = param_mapping[stringer_key][unit_key]
                        cell = get_writable_cell(worksheet, cell_ref)
                        set_cell_value(cell, unit_value)
                        apply_cell_formatting(cell, horizontal='center')
                        print(f"Filled observation {unit_value} for parameter {param_id} at {stringer_key}_{unit_key} in cell {cell_ref}")

def handle_cell_width_parameter(worksheet, value, param_mapping):
    """Handle cell width measurements"""
    if isinstance(value, dict):
        for stringer_key, stringer_value in value.items():
            if stringer_key in param_mapping:
                for position_key, position_value in stringer_value.items():
                    if position_key in param_mapping[stringer_key]:
                        cell_ref = param_mapping[stringer_key][position_key]
                        cell = get_writable_cell(worksheet, cell_ref)
                        set_cell_value(cell, position_value)
                        apply_cell_formatting(cell, horizontal='center')

def handle_groove_length_parameter(worksheet, value, param_mapping):
    """Handle groove length measurements"""
    if isinstance(value, dict):
        for stringer_key, stringer_value in value.items():
            if stringer_key in param_mapping:
                for groove_key, groove_value in stringer_value.items():
                    if groove_key in param_mapping[stringer_key]:
                        cell_ref = param_mapping[stringer_key][groove_key]
                        cell = get_writable_cell(worksheet, cell_ref)
                        set_cell_value(cell, groove_value)
                        apply_cell_formatting(cell, horizontal='center')

def handle_machine_temp_parameter(worksheet, value, param_mapping):
    """Handle machine temperature setup"""
    if isinstance(value, dict):
        for stringer_key, stringer_value in value.items():
            if stringer_key in param_mapping:
                for unit_key, unit_value in stringer_value.items():
                    if unit_key in param_mapping[stringer_key]:
                        for temp_key, temp_value in unit_value.items():
                            if temp_key in param_mapping[stringer_key][unit_key]:
                                cell_ref = param_mapping[stringer_key][unit_key][temp_key]
                                cell = get_writable_cell(worksheet, cell_ref)
                                set_cell_value(cell, temp_value)
                                apply_cell_formatting(cell, horizontal='center')

def handle_light_intensity_parameter(worksheet, value, param_mapping):
    """Handle light intensity and time parameters"""
    if isinstance(value, dict):
        for stringer_key, stringer_value in value.items():
            if stringer_key in param_mapping:
                for unit_key, unit_value in stringer_value.items():
                    if unit_key in param_mapping[stringer_key]:
                        for light_key, light_value in unit_value.items():
                            if light_key in param_mapping[stringer_key][unit_key]:
                                cell_ref = param_mapping[stringer_key][unit_key][light_key]
                                cell = get_writable_cell(worksheet, cell_ref)
                                set_cell_value(cell, light_value)
                                apply_cell_formatting(cell, horizontal='center')

def handle_peel_strength_parameter(worksheet, value, param_mapping):
    """Handle peel strength parameters"""
    if isinstance(value, dict):
        for stringer_key, stringer_value in value.items():
            if stringer_key in param_mapping:
                if 'frontUnit' in param_mapping[stringer_key]:
                    cell_ref = param_mapping[stringer_key]['frontUnit']
                    cell = get_writable_cell(worksheet, cell_ref)
                    set_cell_value(cell, stringer_value.get('frontUnit', ''))
                    apply_cell_formatting(cell, horizontal='center')

                if 'backUnit' in param_mapping[stringer_key]:
                    cell_ref = param_mapping[stringer_key]['backUnit']
                    cell = get_writable_cell(worksheet, cell_ref)
                    set_cell_value(cell, stringer_value.get('backUnit', ''))
                    apply_cell_formatting(cell, horizontal='center')
                front_side = stringer_value.get('frontSide', {})
                for pos_key, pos_value in front_side.items():
                    if pos_key in param_mapping[stringer_key].get('frontSide', {}):
                        cell_ref = param_mapping[stringer_key]['frontSide'][pos_key]
                        cell = get_writable_cell(worksheet, cell_ref)
                        set_cell_value(cell, pos_value)
                        apply_cell_formatting(cell, horizontal='center')
                back_side = stringer_value.get('backSide', {})
                for pos_key, pos_value in back_side.items():
                    if pos_key in param_mapping[stringer_key].get('backSide', {}):
                        cell_ref = param_mapping[stringer_key]['backSide'][pos_key]
                        cell = get_writable_cell(worksheet, cell_ref)
                        set_cell_value(cell, pos_value)
                        apply_cell_formatting(cell, horizontal='center')

def handle_stringer_based_parameter(worksheet, param_id, value, param_mapping):
    """Handle stringer-based parameters"""
    if isinstance(value, dict):
        for stringer_key, stringer_value in value.items():
            if stringer_key in param_mapping:
                if isinstance(stringer_value, dict):
                    # For parameters with time-based values
                    for time_key, time_value in stringer_value.items():
                        if time_key in param_mapping[stringer_key]:
                            cell_ref = param_mapping[stringer_key][time_key]
                            cell = get_writable_cell(worksheet, cell_ref)
                            set_cell_value(cell, time_value)
                            apply_cell_formatting(cell, horizontal='center')
                else:
                    # For simple stringer values
                    cell_ref = param_mapping[stringer_key]
                    cell = get_writable_cell(worksheet, cell_ref)
                    set_cell_value(cell, stringer_value)
                    apply_cell_formatting(cell, horizontal='center')

def handle_bus_ribbon_status(worksheet, value, param_mapping):
    """Handle BUS ribbon status"""
    if isinstance(value, dict):
        normalized_value = dict(value)
        line_prefixes = {key.rsplit('-', 1)[0] for key in param_mapping if key.endswith('-Supplier')}
        for line_prefix in line_prefixes:
            for base_label in ('Width', 'Thickness', 'Expiry Date'):
                top_bottom_key = f"{line_prefix}-{base_label} Top & Bottom"
                middle_key = f"{line_prefix}-{base_label} Middle"
                old_key = f"{line_prefix}-{base_label}"
                if is_empty_value(normalized_value.get(top_bottom_key)) and old_key in normalized_value:
                    normalized_value[top_bottom_key] = normalized_value.get(old_key)
                if top_bottom_key in normalized_value or middle_key in normalized_value:
                    top_bottom_value = normalized_value.get(top_bottom_key, '')
                    middle_value = normalized_value.get(middle_key, '')
                    normalized_value[top_bottom_key] = f"{top_bottom_value} & {middle_value}"
                    normalized_value.pop(middle_key, None)

        for key, val in normalized_value.items():
            if key in param_mapping:
                cell_ref = param_mapping[key]
                cell = get_writable_cell(worksheet, cell_ref)
                set_cell_value(cell, val)
                apply_cell_formatting(cell, horizontal='center')

def handle_soldering_time(worksheet, value, param_mapping):
    """Handle soldering time parameters"""
    if isinstance(value, dict):
        for key, val in value.items():
            if key in param_mapping and isinstance(param_mapping[key], dict) and isinstance(val, dict):
                for field_key, field_value in val.items():
                    if field_key in param_mapping[key]:
                        cell_ref = param_mapping[key][field_key]
                        cell = get_writable_cell(worksheet, cell_ref)
                        set_cell_value(cell, field_value)
                        apply_cell_formatting(cell, horizontal='center')
            elif key in param_mapping:
                cell_ref = param_mapping[key]
                cell = get_writable_cell(worksheet, cell_ref)
                set_cell_value(cell, val)
                apply_cell_formatting(cell, horizontal='center')

def handle_nested_measurements(worksheet, value, param_mapping):
    """Write nested line/sample/measurement structures to their mapped cells."""
    if not isinstance(value, dict) or not isinstance(param_mapping, dict):
        return

    for key, nested_value in value.items():
        nested_mapping = param_mapping.get(key)
        if isinstance(nested_value, dict) and isinstance(nested_mapping, dict):
            handle_nested_measurements(worksheet, nested_value, nested_mapping)
        elif isinstance(nested_mapping, str):
            cell = get_writable_cell(worksheet, nested_mapping)
            set_cell_value(cell, nested_value)
            apply_cell_formatting(cell, horizontal='center')

def handle_line_measurements(worksheet, param_id, value, param_mapping):
    """Handle line-based measurement parameters"""
    if isinstance(value, dict):
        for key, val in value.items():
            if key in param_mapping:
                cell_ref = param_mapping[key]
                cell = get_writable_cell(worksheet, cell_ref)
                set_cell_value(cell, val)
                apply_cell_formatting(cell, horizontal='center')

def handle_bus_peel_strength(worksheet, value, param_mapping):
    """Handle bus ribbon peel strength"""
    if isinstance(value, dict):
        for key, val in value.items():
            if key in param_mapping:
                cell_ref = param_mapping[key]
                cell = get_writable_cell(worksheet, cell_ref)
                set_cell_value(cell, val)
                apply_cell_formatting(cell, horizontal='center')

def handle_time_based_line_parameter(worksheet, param_id, value, param_mapping, safety_module_ids=None):
    """Handle time-based line parameters"""
    if isinstance(value, dict):
        safety_module_ids = safety_module_ids or {}
        for key, val in value.items():
            if param_id == '26-1' and is_safety_module_id_field(key):
                continue
            mapping_key = get_safety_mapping_key(key, param_mapping) if param_id in SAFETY_TEST_PARAM_IDS else key
            if mapping_key in param_mapping:
                cell_ref = param_mapping[mapping_key]
                cell = get_writable_cell(worksheet, cell_ref)
                if param_id in SAFETY_TEST_PARAM_IDS:
                    source = value if param_id == '26-1' else safety_module_ids
                    module_id = get_safety_module_id_for_result(source, key)
                    val = combine_result_with_module_id(val, module_id)
                set_cell_value(cell, val)
                apply_cell_formatting(cell, horizontal='center')

def handle_component_status(worksheet, param_id, value, param_mapping):
    """Handle component status parameters"""
    if isinstance(value, dict):
        for key, val in value.items():
            if key in param_mapping:
                cell_ref = param_mapping[key]
                cell = get_writable_cell(worksheet, cell_ref)
                set_cell_value(cell, val)
                apply_cell_formatting(cell, horizontal='center')

def handle_laminator_parameter(worksheet, param_id, value, param_mapping):
    """Handle laminator parameters"""
    if isinstance(value, dict):
        # Handle upper chamber
        upper_data = value.get('upper', {})
        for key, val in upper_data.items():
            if key in param_mapping.get('upper', {}):
                cell_ref = param_mapping['upper'][key]
                cell = get_writable_cell(worksheet, cell_ref)
                set_cell_value(cell, val)
                apply_cell_formatting(cell, horizontal='center')

        # Handle lower chamber
        lower_data = value.get('lower', {})
        for key, val in lower_data.items():
            if key in param_mapping.get('lower', {}):
                cell_ref = param_mapping['lower'][key]
                cell = get_writable_cell(worksheet, cell_ref)
                set_cell_value(cell, val)
                apply_cell_formatting(cell, horizontal='center')

        # Handle selected recipe
        if 'selectedRecipe' in param_mapping:
            cell_ref = param_mapping['selectedRecipe']
            cell = get_writable_cell(worksheet, cell_ref)
            set_cell_value(cell, value.get('selectedRecipe', ''))
            apply_cell_formatting(cell, horizontal='center')

def handle_sample_line_parameter(worksheet, param_id, value, param_mapping):
    """Handle sample-based line parameters"""
    if isinstance(value, dict):
        for key, val in value.items():
            if key in param_mapping:
                cell_ref = param_mapping[key]
                cell = get_writable_cell(worksheet, cell_ref)
                set_cell_value(cell, val)
                apply_cell_formatting(cell, horizontal='center')

def handle_frame_sealant_weight(worksheet, value, param_mapping):
    """Handle frame sealant weight"""
    if isinstance(value, dict):
        for key, val in value.items():
            if key in param_mapping:
                cell_ref = param_mapping[key]
                cell = get_writable_cell(worksheet, cell_ref)
                set_cell_value(cell, val)
                apply_cell_formatting(cell, horizontal='center')

def handle_jb_measurements(worksheet, param_id, value, param_mapping):
    """Handle junction box measurements"""
    if isinstance(value, dict):
        for key, val in value.items():
            if key in param_mapping:
                cell_ref = param_mapping[key]
                cell = get_writable_cell(worksheet, cell_ref)
                set_cell_value(cell, val)
                apply_cell_formatting(cell, horizontal='center')

def handle_complex_measurements(worksheet, param_id, value, param_mapping):
    """Handle complex measurement parameters"""
    if isinstance(value, dict):
        for key, val in value.items():
            if key in param_mapping:
                cell_ref = param_mapping[key]
                cell = get_writable_cell(worksheet, cell_ref)
                set_cell_value(cell, val)
                apply_cell_formatting(cell, horizontal='center')

def handle_environmental_parameters(worksheet, param_id, value, param_mapping):
    """Handle environmental parameters"""
    if isinstance(value, dict):
        for key, val in value.items():
            if key in param_mapping:
                cell_ref = param_mapping[key]
                cell = get_writable_cell(worksheet, cell_ref)
                set_cell_value(cell, val)
                apply_cell_formatting(cell, horizontal='center')

def adjust_column_widths(worksheet, min_width=8, max_width=50):
    """
    Automatically adjust column widths based on content
    """
    for column in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        
        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        
        adjusted_width = min(max_length + 2, max_width)
        adjusted_width = max(adjusted_width, min_width)
        
        worksheet.column_dimensions[column_letter].width = adjusted_width

def generate_filename(audit_data):
    line_number = audit_data.get('lineNumber', 'Unknown')
    date = audit_data.get('date', 'Unknown')
    shift = audit_data.get('shift', 'Unknown')
    formatted_date = date.replace('-', '')
    return f"Quality_Audit_Line{line_number}_{formatted_date}_Shift{shift}.xlsx"

# Main function to generate report (this will be called from main.py)
def generate_audit_report(audit_data):
    try:
        if not audit_data:
            raise ValueError("No audit data provided")
        audit_data = normalize_ipqc_audit_data(copy.deepcopy(audit_data))
        print("Received audit data for report generation")
        line_number = audit_data.get('lineNumber', 'I')
        template_path, field_config, observation_cell_mapping = get_template_config(line_number)
        if not template_path or not os.path.exists(template_path):
            raise FileNotFoundError(f"Audit template could not be loaded for line {line_number}: {template_path}")
        if not isinstance(observation_cell_mapping, dict):
            print("Warning: observation cell mapping is invalid; continuing without observations")
            observation_cell_mapping = {}
        wb = load_workbook(template_path)
        setup_cell_styles(wb)
        ws = wb.active
        fill_basic_info(ws, audit_data, field_config)
        fill_observations_data(ws, audit_data, observation_cell_mapping)
        fill_line_dropdown_values(ws, audit_data)
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        filename = generate_filename(audit_data)
        return output, filename
        
    except Exception as e:
        print(f"Error generating audit report: {str(e)}")
        raise
