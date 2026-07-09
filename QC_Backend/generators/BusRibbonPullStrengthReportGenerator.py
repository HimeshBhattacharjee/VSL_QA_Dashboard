from logging_utils import log_progress
import calendar
from datetime import datetime
from functools import lru_cache
import io
import os
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from generators.excel_image_utils import add_worksheet_images, collect_worksheet_images
from paths import get_template_key
from s3_service import get_s3_client

LINE_BUSSING_KEYS = {
    "FAB-II Line-I": ("autoBussing1", "autoBussing2", "autoBussing3"),
    "FAB-II Line-II": ("autoBussing4", "autoBussing5"),
}
BUSSING_LABELS = {
    "autoBussing1": 1,
    "autoBussing2": 2,
    "autoBussing3": 3,
    "autoBussing4": 4,
    "autoBussing5": 5,
}
SHIFT_ROW_RANGES = {
    "A": (7, 12),
    "B": (13, 18),
    "C": (19, 24),
}
STRENGTH_START_COLUMN = 9
AVERAGE_COLUMN = 25
TEMPLATE_FILENAME = "Blank Bus Ribbon to INTC Ribbon Pull Test Report.xlsx"
MIN_STRENGTH_N = 1.5
LOW_STRENGTH_FILL = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
LOW_STRENGTH_FONT = Font(name='Times New Roman', size=10, color="C00000")

def normalize_line(line):
    return "FAB-II Line-II" if line == "FAB-II Line-II" else "FAB-II Line-I"

def safe_filename(value):
    clean_name = "".join(c for c in str(value or "") if c.isalnum() or c in (" ", "-", "_")).strip()
    return clean_name.replace(" ", "_") or "Bus_Ribbon_INTC_Pull_Strength"

def parse_numeric(value):
    if value in (None, ""):
        return None
    try:
        return float(str(value).strip())
    except (TypeError, ValueError):
        return None

def to_excel_number(value):
    numeric_value = parse_numeric(value)
    if numeric_value is None:
        return "" if value is None else value
    return int(numeric_value) if numeric_value.is_integer() else numeric_value

def calculate_average(values):
    numeric_values = [value for value in (parse_numeric(item) for item in values) if value is not None]
    if not numeric_values:
        return ""
    return round(sum(numeric_values) / len(numeric_values), 2)

def is_strength_below_limit(value):
    numeric_value = parse_numeric(value)
    return numeric_value is not None and numeric_value < MIN_STRENGTH_N

@lru_cache(maxsize=1)
def get_template_bytes():
    s3_client = get_s3_client()
    bucket = os.getenv("S3_BUCKET_NAME")
    if not bucket:
        raise ValueError("S3_BUCKET_NAME environment variable not set")
    response = s3_client.get_object(Bucket=bucket, Key=get_template_key(TEMPLATE_FILENAME))
    return response["Body"].read()

def load_template_workbook():
    return load_workbook(io.BytesIO(get_template_bytes()))

def get_entry_signatures(entries):
    for entry in entries:
        signatures = entry.get("signatures") or {}
        if signatures.get("preparedBy") or signatures.get("reviewedBy") or signatures.get("verifiedBy"):
            return signatures
    return {}

def write_strength_cell(worksheet, row, column, value):
    cell = worksheet.cell(row=row, column=column, value=to_excel_number(value))
    if is_strength_below_limit(value):
        cell.fill = LOW_STRENGTH_FILL
        cell.font = LOW_STRENGTH_FONT

def write_machine_rows(worksheet, entry, machine_key, top_row):
    shift_details = entry.get("shiftDetails", {}) or {}
    bussing_data = entry.get("bussingData", {}) or {}
    averages = entry.get("averages", {}) or {}
    machine = bussing_data.get(machine_key, {}) or {}
    strengths = machine.get("strengths", []) or []
    machine_averages = averages.get(machine_key, {}) or {}
    worksheet[f"D{top_row}"] = shift_details.get("poNumber", "")
    worksheet[f"E{top_row}"] = shift_details.get("intcRibbonStatus", "")
    worksheet[f"F{top_row}"] = shift_details.get("busRibbonStatus", "")
    worksheet[f"G{top_row}"] = BUSSING_LABELS[machine_key]
    worksheet[f"H{top_row}"] = machine.get("position", "")
    first_strength_row = top_row
    second_strength_row = top_row + 1
    for index in range(16):
        write_strength_cell(
            worksheet,
            first_strength_row,
            STRENGTH_START_COLUMN + index,
            strengths[index] if index < len(strengths) else "",
        )
        write_strength_cell(
            worksheet,
            second_strength_row,
            STRENGTH_START_COLUMN + index,
            strengths[index + 16] if index + 16 < len(strengths) else "",
        )
    worksheet.cell(
        row=first_strength_row,
        column=AVERAGE_COLUMN,
        value=to_excel_number(machine_averages.get("average1") or calculate_average(strengths[:16])),
    )
    worksheet.cell(
        row=second_strength_row,
        column=AVERAGE_COLUMN,
        value=to_excel_number(machine_averages.get("average2") or calculate_average(strengths[16:32])),
    )

def fill_bus_ribbon_data_in_sheet(worksheet, entries, date_label, line):
    shift_order = {"A": 0, "B": 1, "C": 2}
    sorted_entries = sorted(entries, key=lambda item: shift_order.get(item.get("shift", ""), 9))
    entry_by_shift = {entry.get("shift"): entry for entry in sorted_entries if entry.get("shift") in SHIFT_ROW_RANGES}
    for shift, (start_row, _) in SHIFT_ROW_RANGES.items():
        entry = entry_by_shift.get(shift)
        if not entry:
            continue
        worksheet[f"B{start_row}"] = date_label
        worksheet[f"C{start_row}"] = shift
        for machine_index, machine_key in enumerate(LINE_BUSSING_KEYS[line]):
            write_machine_rows(worksheet, entry, machine_key, start_row + (machine_index * 2))
    signatures = get_entry_signatures(sorted_entries)
    if signatures.get("preparedBy"):
        worksheet["E25"] = signatures.get("preparedBy", "")
    if signatures.get("reviewedBy") or signatures.get("verifiedBy"):
        worksheet["Q25"] = signatures.get("reviewedBy") or signatures.get("verifiedBy", "")

def generate_bus_ribbon_pull_strength_report(report_data):
    try:
        if not report_data:
            raise ValueError("No bus ribbon pull strength data provided")
        if isinstance(report_data, list):
            entries = report_data
            year = datetime.now().year
            month = datetime.now().month
            line = "FAB-II Line-I"
        else:
            entries = report_data.get("entries", [])
            year = int(report_data.get("year", datetime.now().year))
            month = int(report_data.get("month", datetime.now().month))
            line = normalize_line(report_data.get("line"))
        entries_by_date = {}
        for entry in entries:
            if normalize_line(entry.get("line")) != line:
                continue
            date = str(entry.get("date", "")).split("T")[0]
            if date:
                entries_by_date.setdefault(date, []).append(entry)
        workbook = load_template_workbook()
        template_sheet = workbook.active
        template_images = collect_worksheet_images(template_sheet)

        days_in_month = calendar.monthrange(year, month)[1]
        for day in range(1, days_in_month + 1):
            date_key = f"{year}-{month:02d}-{day:02d}"
            date_label = f"{day:02d}.{month:02d}.{year}"
            worksheet = workbook.copy_worksheet(template_sheet)
            worksheet.title = date_label
            add_worksheet_images(worksheet, template_images)
            date_entries = entries_by_date.get(date_key, [])
            if date_entries:
                fill_bus_ribbon_data_in_sheet(worksheet, date_entries, date_label, line)
        if len(workbook.worksheets) > 1:
            workbook.remove(template_sheet)
        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)
        month_name = datetime(year, month, 1).strftime("%B")
        filename = f"{safe_filename('Bus_Ribbon_INTC_Pull_Strength')}_{safe_filename(line)}_{month_name}_{year}.xlsx"
        return output, filename
    except Exception as e:
        log_progress(f"Error generating bus ribbon pull strength report: {str(e)}")
        raise
