from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
import io
from datetime import datetime
import calendar
from paths import get_template_key, download_from_s3
from excel_image_utils import copy_worksheet_images

GLASS_GROOVE_TARGETS = {
    'Glass Groove (5.6 mm)': 40,
    'Glass Groove (6.1 mm)': 49
}

FRAME_SEALANT_PASS_TOLERANCE = 7
PASS_FILL = PatternFill(start_color='92D050', end_color='92D050', fill_type='solid')
FAIL_FILL = PatternFill(start_color='FF9999', end_color='FF9999', fill_type='solid')
PASS_FONT = Font(color='006100')
FAIL_FONT = Font(color='9C0006')

def get_display_line_numbers(line_group):
    return ('3', '4') if line_group == 'Line-II' else ('1', '2')

def apply_line_weight_format(cell, glass_groove, weight_per_meter):
    target_weight = GLASS_GROOVE_TARGETS.get(str(glass_groove).strip())
    if target_weight is None or weight_per_meter in (None, ''):
        return

    try:
        weight_value = float(weight_per_meter)
    except (ValueError, TypeError):
        return

    min_weight = target_weight - FRAME_SEALANT_PASS_TOLERANCE
    max_weight = target_weight + FRAME_SEALANT_PASS_TOLERANCE

    if min_weight <= weight_value <= max_weight:
        cell.fill = PASS_FILL
        cell.font = PASS_FONT
    else:
        cell.fill = FAIL_FILL
        cell.font = FAIL_FONT

def fill_frame_data_in_sheet(worksheet, entries, date):
    """Fill frame sealant weight data for a specific date into a worksheet"""
    try:
        # Sort entries by shift
        shift_order = {'A': 0, 'B': 1, 'C': 2}
        sorted_entries = sorted(entries, key=lambda e: shift_order.get(e.get('shift', ''), 3))
        
        # Define row ranges for each shift
        shift_rows = {
            'A': {'start': 7, 'end': 10},    # A shift rows 7-10
            'B': {'start': 11, 'end': 14},   # B shift rows 11-14
            'C': {'start': 15, 'end': 18}    # C shift rows 15-18
        }
        
        # Create a dictionary to map shift to entry
        entry_by_shift = {}
        for entry in sorted_entries:
            shift = entry.get('shift', '')
            if shift:
                entry_by_shift[shift] = entry
        
        # Process all shifts in order
        for shift in ['A', 'B', 'C']:
            entry = entry_by_shift.get(shift)
            
            if entry:
                # Entry exists for this shift, fill data
                current_row = shift_rows[shift]['start']
                fill_shift_data(worksheet, entry, date, current_row)
            else:
                # No entry for this shift, leave rows empty
                # Optionally, you can clear the rows or leave them as is
                # The rows will remain empty from the template
                pass
        
        # Fill signatures at the bottom (using the first entry if available)
        if sorted_entries and sorted_entries[0].get('signatures'):
            signatures = sorted_entries[0]['signatures']
            if signatures.get('preparedBy'):
                worksheet['E19'] = signatures.get('preparedBy', '')
            if signatures.get('verifiedBy'):
                worksheet['M19'] = signatures.get('verifiedBy', '')
        
        print(f"Filled frame sealant data for date {date}")
        
    except Exception as e:
        print(f"Error filling frame sealant data in sheet: {str(e)}")
        raise


def fill_shift_data(worksheet, entry, date, current_row):
    """Helper function to fill data for a single shift at the specified starting row"""
    shift = entry.get('shift', '')
    lines = entry.get('lines', {})
    display_line_1, display_line_2 = get_display_line_numbers(entry.get('lineGroup'))
    
    def write_line_rows(line_number, line_data, base_row):
        po_number = line_data.get('po', '')
        length_data = line_data.get('length', {})
        width_data = line_data.get('width', {})

        row_mappings = [
            (base_row, f"Length - {length_data.get('frameSize', '')}", length_data),
            (base_row + 1, f"Width - {width_data.get('frameSize', '')}", width_data),
        ]

        for row_idx, frame_label, division_data in row_mappings:
            worksheet[f'A{row_idx}'] = date
            worksheet[f'B{row_idx}'] = f"Shift {shift}"
            worksheet[f'C{row_idx}'] = line_number
            worksheet[f'D{row_idx}'] = po_number
            worksheet[f'E{row_idx}'] = division_data.get('frameSupplier', '')
            worksheet[f'F{row_idx}'] = frame_label
            worksheet[f'G{row_idx}'] = division_data.get('sealantSupplier', '')
            worksheet[f'H{row_idx}'] = division_data.get('sealantExpiry', '')
            worksheet[f'I{row_idx}'] = division_data.get('frameWithoutSealant1', '')
            worksheet[f'J{row_idx}'] = division_data.get('frameWithoutSealant2', '')
            worksheet[f'K{row_idx}'] = division_data.get('frameWithSealant1', '')
            worksheet[f'L{row_idx}'] = division_data.get('frameWithSealant2', '')
            worksheet[f'M{row_idx}'] = division_data.get('netSealantWeight1', '')
            worksheet[f'N{row_idx}'] = division_data.get('netSealantWeight2', '')

        worksheet[f'O{base_row}'] = line_data.get('totalSealantWeightPerModule', '')
        worksheet[f'P{base_row}'] = line_data.get('sealantWeightPerModulePerMeter', '')
        worksheet[f'Q{base_row}'] = line_data.get('remarks', '')
        apply_line_weight_format(
            worksheet[f'P{base_row}'],
            line_data.get('glassGroove', ''),
            line_data.get('sealantWeightPerModulePerMeter', '')
        )

    write_line_rows(display_line_1, lines.get('1', {}), current_row)
    write_line_rows(display_line_2, lines.get('2', {}), current_row + 2)

def generate_frame_sealant_report(frame_data):
    """Generate frame sealant weight report with multiple sheets - one sheet per day of the month"""
    try:
        if not frame_data:
            raise ValueError("No frame sealant data provided")
        
        print("Received frame sealant data for report generation")
        
        # Handle different input formats
        if isinstance(frame_data, list):
            entries = frame_data
            year = datetime.now().year
            month = datetime.now().month
        else:
            entries = frame_data.get('entries', [])
            year = frame_data.get('year', datetime.now().year)
            month = frame_data.get('month', datetime.now().month)
        
        print(f"Entries count: {len(entries)}")
        print(f"Year: {year}, Month: {month}")
        
        # Group entries by date
        entries_by_date = {}
        for entry in entries:
            date = entry.get('date', '')
            if date not in entries_by_date:
                entries_by_date[date] = []
            entries_by_date[date].append(entry)
        
        # Get days in month
        days_in_month = calendar.monthrange(year, month)[1]
        
        # Load template
        template_key = get_template_key('Blank Frame Sealant Weight Report.xlsx')
        template_path = download_from_s3(template_key)
        
        # Load workbook
        wb = load_workbook(template_path)
        
        template_sheet = wb.active

        # Create sheets for each day of the month
        for day in range(1, days_in_month + 1):
            date_str = f"{year}-{month:02d}-{day:02d}"
            formatted_date = f"{day:02d}.{month:02d}.{year}"
            
            # Create new sheet by copying template
            new_sheet = wb.copy_worksheet(template_sheet)
            new_sheet.title = formatted_date
            copy_worksheet_images(template_sheet, new_sheet)
            
            # Get entries for this date
            date_entries = entries_by_date.get(date_str, [])
            
            if date_entries:
                # Fill data if entries exist
                fill_frame_data_in_sheet(new_sheet, date_entries, formatted_date)
        
        # Remove the original template sheet (first sheet)
        if len(wb.worksheets) > 1:
            wb.remove(template_sheet)
        
        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Generate filename
        month_name = datetime(year, month, 1).strftime("%B")
        filename = f"Frame_Sealant_Weight_{month_name}_{year}.xlsx"
        
        print(f"Frame sealant report generated successfully: {filename}")
        return output, filename
        
    except Exception as e:
        print(f"Error generating frame sealant report: {str(e)}")
        raise
