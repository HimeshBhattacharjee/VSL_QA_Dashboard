from openpyxl import load_workbook
from openpyxl.styles import Alignment
import io
from paths import get_template_key, download_from_s3

def fill_allowable_limit_with_checkboxes(worksheet, gel_data):
    try:
        form_data = gel_data.get('form_data', {})
        checkbox_0 = form_data.get('checkbox_0', False)
        checkbox_1 = form_data.get('checkbox_1', False)
        eva_epe_symbol = "✓" if checkbox_0 else "✕"
        poe_symbol = "✓" if checkbox_1 else "✕"
        allowable_limit_EVA = (f"1. Gel Content should be: 75 to 95% for EVA & EPE {eva_epe_symbol}")
        worksheet['B6'] = allowable_limit_EVA
        worksheet['B6'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        allowable_limit_POE = (f"2. Gel Content should be: ≥ 60% for POE {poe_symbol}")
        worksheet['B7'] = allowable_limit_POE
        worksheet['B7'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        print("Allowable limit with checkboxes filled successfully in cell B6 and B7")
    except Exception as e:
        print(f"Error filling allowable limit with checkboxes: {str(e)}")
        raise

def fill_encapsulant_types_with_checkboxes(worksheet, gel_data):
    try:
        form_data = gel_data.get('form_data', {})
        checkbox_2 = form_data.get('checkbox_2', False)
        checkbox_3 = form_data.get('checkbox_3', False)
        checkbox_4 = form_data.get('checkbox_4', False)
        eva_symbol = "✓" if checkbox_2 else "✕"
        epe_symbol = "✓" if checkbox_3 else "✕"
        poe_symbol = "✓" if checkbox_4 else "✕"
        encapsulant_text = f"EVA {eva_symbol}           EPE {epe_symbol}           POE {poe_symbol}"
        worksheet['I10'] = encapsulant_text
        worksheet['I10'].alignment = Alignment(horizontal='center', vertical='center')
        print("Encapsulant types with checkboxes filled successfully in cell I10")
    except Exception as e:
        print(f"Error filling encapsulant types with checkboxes: {str(e)}")
        raise

def fill_gel_basic_info(worksheet, gel_data):
    try:
        form_data = gel_data.get('form_data', {})
        basic_info_mapping = {
            'gel_editable_0': 'I5',
            'gel_editable_1': 'I6',
            'gel_editable_2': 'I7',
            'gel_editable_3': 'I8',
            'gel_editable_4': 'C10',
            'gel_editable_5': 'D10',
            'gel_editable_6': 'E10',
            'gel_editable_7': 'C11',
            'gel_editable_8': 'D11',
            'gel_editable_9': 'E11',
            'gel_editable_10': 'C12',
            'gel_editable_11': 'D12',
            'gel_editable_12': 'E12',
            'gel_editable_13': 'I12',
            'gel_editable_14': 'J12',
            'gel_editable_15': 'K12',
            'gel_editable_16': 'L12',
            'gel_editable_17': 'C13',
            'gel_editable_18': 'D13',
            'gel_editable_19': 'E13',
            'gel_editable_20': 'I13',
            'gel_editable_21': 'J13',
            'gel_editable_22': 'K13',
            'gel_editable_23': 'L13',
            'gel_editable_24': 'C14',
            'gel_editable_25': 'D14',
            'gel_editable_26': 'E14',
            'gel_editable_27': 'I14',
            'gel_editable_28': 'J14',
            'gel_editable_29': 'K14',
            'gel_editable_30': 'L14',
            'gel_editable_31': 'C15',
            'gel_editable_32': 'D15',
            'gel_editable_33': 'E15',
            'gel_editable_34': 'I15',
            'gel_editable_35': 'J15',
            'gel_editable_36': 'K15',
            'gel_editable_37': 'L15',
            'gel_editable_38': 'C16',
            'gel_editable_39': 'D16',
            'gel_editable_40': 'E16',
            'gel_editable_41': 'I16',
            'gel_editable_42': 'B19',
            'gel_editable_53': 'B21',
            'gel_editable_69': 'B23',
            'preparedBySignature': 'C27',
            'acceptedBySignature': 'H27'
        }
        for field, cell_ref in basic_info_mapping.items():
            if field in form_data and form_data[field]:
                worksheet[cell_ref] = form_data[field]
        fill_allowable_limit_with_checkboxes(worksheet, gel_data)
        fill_encapsulant_types_with_checkboxes(worksheet, gel_data)
        print("Basic gel test information filled successfully")
    except Exception as e:
        print(f"Error filling basic gel test info: {str(e)}")
        raise

def fill_gel_test_data(worksheet, gel_data):
    try:
        form_data = gel_data.get('form_data', {})
        averages = gel_data.get('averages', {})
        test_data_mapping = {
            'gel_editable_43': 'E19', 'gel_editable_44': 'F19', 'gel_editable_45': 'G19', 'gel_editable_46': 'H19', 'gel_editable_47': 'I19', 
            'gel_editable_48': 'E20', 'gel_editable_49': 'F20', 'gel_editable_50': 'G20', 'gel_editable_51': 'H20', 'gel_editable_52': 'I20',
            'gel_editable_54': 'E21', 'gel_editable_55': 'F21', 'gel_editable_56': 'G21', 'gel_editable_57': 'H21', 'gel_editable_58': 'I21',
            'gel_editable_59': 'E22', 'gel_editable_60': 'F22', 'gel_editable_61': 'G22', 'gel_editable_62': 'H22', 'gel_editable_63': 'I22',
            'gel_editable_64': 'E23', 'gel_editable_65': 'F23', 'gel_editable_66': 'G23', 'gel_editable_67': 'H23', 'gel_editable_68': 'I23',
            'gel_editable_70': 'E24', 'gel_editable_71': 'F24', 'gel_editable_72': 'G24', 'gel_editable_73': 'H24', 'gel_editable_74': 'I24',
            'gel_editable_75': 'E25', 'gel_editable_76': 'F25', 'gel_editable_77': 'G25', 'gel_editable_78': 'H25', 'gel_editable_79': 'I25'
        }
        for field, cell_ref in test_data_mapping.items():
            if field in form_data and form_data[field]:
                worksheet[cell_ref] = form_data[field]
        averages_mapping = {
            'average_0': 'J19',
            'average_1': 'J20',
            'average_2': 'J21',
            'average_3': 'J22',
            'average_4': 'J23',
            'average_5': 'J24',
            'average_6': 'J25' 
        }
        for field, cell_ref in averages_mapping.items():
            if field in averages and averages[field]:
                worksheet[cell_ref] = averages[field]
        if 'mean' in averages and averages['mean']:
            worksheet.merge_cells('L19:L25')
            worksheet['L19'] = averages['mean']
        print("Gel test data filled successfully")
    except Exception as e:
        print(f"Error filling gel test data: {str(e)}")
        raise

def generate_gel_filename(gel_data):
    report_name = gel_data.get('report_name', 'Gel_Test_Report')
    timestamp = gel_data.get('timestamp', '').split('T')[0] if gel_data.get('timestamp') else ''
    clean_name = "".join(c for c in report_name if c.isalnum() or c in (' ', '-', '_')).rstrip()
    clean_name = clean_name.replace(' ', '_')
    if timestamp:
        return f"Gel_Test_{clean_name}_{timestamp}.xlsx"
    else:
        return f"Gel_Test_{clean_name}.xlsx"

def generate_gel_report(gel_data):
    try:
        if not gel_data:
            raise ValueError("No gel test data provided")
        print("Received gel test data for report generation")
        print(f"Report name: {gel_data.get('report_name', 'N/A')}")
        print(f"Form data keys: {list(gel_data.get('form_data', {}).keys())}")
        print(f"Averages keys: {list(gel_data.get('averages', {}).keys())}")
        template_key = get_template_key('Blank Gel Content Test Report.xlsx')
        template_path = download_from_s3(template_key)
        wb = load_workbook(template_path)
        ws = wb.active
        fill_gel_basic_info(ws, gel_data)
        fill_gel_test_data(ws, gel_data)
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        filename = generate_gel_filename(gel_data)
        print(f"Gel test report generated successfully: {filename}")
        print("Text with inline checkboxes have been added to single cells")
        return output, filename
    except Exception as e:
        print(f"Error generating gel test report: {str(e)}")
        raise