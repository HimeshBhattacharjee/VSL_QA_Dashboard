from logging_utils import log_progress
import calendar
from copy import copy
from datetime import datetime
from functools import lru_cache
import io
import os
import re

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill

from generators.excel_image_utils import add_worksheet_images, collect_worksheet_images
from paths import get_template_key
from s3_service import get_s3_client


TEMPLATE_FILENAME = "Blank JB Contact Block Maintenance Report.xlsx"
FAB_LINES = {
    "FAB-II Line-I": ("Line - 1", "Line - 2"),
    "FAB-II Line-II": ("Line - 3", "Line - 4"),
}
DATA_START_ROW = 6
SIGNATURE_BLOCK_START_ROW = 56
BUILT_IN_DATA_ROWS = 50
STYLE_TEMPLATE_ROW = 55
ROW_COLUMNS = {
    "date": "C",
    "line": "D",
    "po": "E",
    "jbNo": "F",
    "sortValuePositive": "G",
    "sortValueNegative": "H",
    "springTension": "I",
    "remarks": "J",
    "checkedBy": "K",
}
NUMERIC_FIELDS = ("sortValuePositive", "sortValueNegative", "springTension")
NUMERIC_ONLY_RE = re.compile(r"^[+-]?(?:\d+(?:\.\d*)?|\.\d+)$")
FAIL_FILL = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
FAIL_FONT = Font(color="C00000")


def normalize_fab(fab):
    return "FAB-II Line-II" if fab == "FAB-II Line-II" else "FAB-II Line-I"


def safe_filename(value):
    clean_name = "".join(c for c in str(value or "") if c.isalnum() or c in (" ", "-", "_")).strip()
    return clean_name.replace(" ", "_") or "JB_Contact_Block_Maintenance"


def parse_numeric(value):
    if value in (None, ""):
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip()
    if not text or not NUMERIC_ONLY_RE.fullmatch(text):
        return None
    try:
        return float(text)
    except (TypeError, ValueError):
        return None


def to_excel_number(value):
    numeric_value = parse_numeric(value)
    if numeric_value is None:
        return ""
    return int(numeric_value) if numeric_value.is_integer() else numeric_value


def calculate_remarks(row):
    spring_tension = parse_numeric(row.get("springTension"))
    sort_value_positive = parse_numeric(row.get("sortValuePositive"))
    sort_value_negative = parse_numeric(row.get("sortValueNegative"))

    if spring_tension is None or sort_value_positive is None or sort_value_negative is None:
        return ""

    return "OK" if spring_tension >= 75 and sort_value_positive < 20 and sort_value_negative < 20 else "NOT OK"


@lru_cache(maxsize=1)
def get_template_bytes():
    s3_client = get_s3_client()
    bucket = os.getenv("S3_BUCKET_NAME")
    if not bucket:
        raise ValueError("S3_BUCKET_NAME environment variable not set")

    response = s3_client.get_object(Bucket=bucket, Key=get_template_key(TEMPLATE_FILENAME))
    return response["Body"].read()


def load_template_workbook():
    return load_workbook(io.BytesIO(get_template_bytes()))


def shift_merged_range(range_boundaries, row_offset, insert_at):
    min_col, min_row, max_col, max_row = range_boundaries
    if min_row >= insert_at:
        min_row += row_offset
        max_row += row_offset
    elif min_row < insert_at <= max_row:
        max_row += row_offset
    return min_col, min_row, max_col, max_row


def ensure_data_rows(worksheet, required_rows):
    extra_rows = max(0, required_rows - BUILT_IN_DATA_ROWS)
    if extra_rows <= 0:
        return

    merged_ranges = [merged_range.bounds for merged_range in worksheet.merged_cells.ranges]
    for merged_range in list(worksheet.merged_cells.ranges):
        worksheet.unmerge_cells(str(merged_range))

    worksheet.insert_rows(SIGNATURE_BLOCK_START_ROW, extra_rows)

    for row_index in range(SIGNATURE_BLOCK_START_ROW, SIGNATURE_BLOCK_START_ROW + extra_rows):
        worksheet.row_dimensions[row_index].height = worksheet.row_dimensions[STYLE_TEMPLATE_ROW].height
        for column_index in range(3, 12):
            source = worksheet.cell(row=STYLE_TEMPLATE_ROW, column=column_index)
            target = worksheet.cell(row=row_index, column=column_index)
            target._style = copy(source._style)
            target.font = copy(source.font)
            target.fill = copy(source.fill)
            target.border = copy(source.border)
            target.alignment = copy(source.alignment)
            target.number_format = source.number_format
            target.protection = copy(source.protection)

    for min_col, min_row, max_col, max_row in (
        shift_merged_range(bounds, extra_rows, SIGNATURE_BLOCK_START_ROW)
        for bounds in merged_ranges
    ):
        worksheet.merge_cells(
            start_row=min_row,
            start_column=min_col,
            end_row=max_row,
            end_column=max_col,
        )


def get_flat_rows(entry, fab):
    flat_rows = []
    lines = entry.get("lines") or {}
    for line_label in FAB_LINES[normalize_fab(fab)]:
        rows = lines.get(line_label) or []
        if not isinstance(rows, list):
            continue
        for row in rows:
            flat_rows.append((line_label, row or {}))
    return flat_rows


def write_data_row(worksheet, row_index, date_label, line_label, row_data):
    worksheet[f"{ROW_COLUMNS['date']}{row_index}"] = date_label
    worksheet[f"{ROW_COLUMNS['line']}{row_index}"] = line_label[len(line_label) - 1]
    worksheet[f"{ROW_COLUMNS['po']}{row_index}"] = row_data.get("po", "")
    worksheet[f"{ROW_COLUMNS['jbNo']}{row_index}"] = row_data.get("jbNo", "")
    worksheet[f"{ROW_COLUMNS['sortValuePositive']}{row_index}"] = to_excel_number(row_data.get("sortValuePositive"))
    worksheet[f"{ROW_COLUMNS['sortValueNegative']}{row_index}"] = to_excel_number(row_data.get("sortValueNegative"))
    worksheet[f"{ROW_COLUMNS['springTension']}{row_index}"] = to_excel_number(row_data.get("springTension"))
    remarks = calculate_remarks(row_data)
    remarks_cell = worksheet[f"{ROW_COLUMNS['remarks']}{row_index}"]
    remarks_cell.value = remarks
    if remarks == "NOT OK":
        remarks_cell.fill = FAIL_FILL
        remarks_cell.font = FAIL_FONT
    worksheet[f"{ROW_COLUMNS['checkedBy']}{row_index}"] = row_data.get("checkedBy", "")


def fill_jb_contact_block_data_in_sheet(worksheet, entry, date_label, fab):
    flat_rows = get_flat_rows(entry, fab)
    ensure_data_rows(worksheet, len(flat_rows))

    for offset, (line_label, row_data) in enumerate(flat_rows):
        write_data_row(worksheet, DATA_START_ROW + offset, date_label, line_label, row_data)

    signature_row = SIGNATURE_BLOCK_START_ROW + max(0, len(flat_rows) - BUILT_IN_DATA_ROWS) + 1
    signatures = entry.get("signatures") or {}
    if signatures.get("preparedBy"):
        worksheet[f"E{signature_row}"] = signatures.get("preparedBy", "")
    if signatures.get("verifiedBy") or signatures.get("reviewedBy"):
        worksheet[f"J{signature_row}"] = signatures.get("verifiedBy") or signatures.get("reviewedBy", "")


def generate_jb_contact_block_maintenance_report(report_data):
    try:
        if not report_data:
            raise ValueError("No JB contact block maintenance data provided")

        if isinstance(report_data, list):
            entries = report_data
            year = datetime.now().year
            month = datetime.now().month
            fab = "FAB-II Line-I"
        else:
            entries = report_data.get("entries", [])
            year = int(report_data.get("year", datetime.now().year))
            month = int(report_data.get("month", datetime.now().month))
            fab = normalize_fab(report_data.get("fab"))

        entry_by_date = {}
        for entry in entries:
            if normalize_fab(entry.get("fab")) != fab:
                continue
            date = str(entry.get("date", "")).split("T")[0]
            if date:
                entry_by_date[date] = entry

        workbook = load_template_workbook()
        template_sheet = workbook.active
        template_images = collect_worksheet_images(template_sheet)

        days_in_month = calendar.monthrange(year, month)[1]
        for day in range(1, days_in_month + 1):
            date_key = f"{year}-{month:02d}-{day:02d}"
            date_label = f"{day:02d}.{month:02d}.{year}"
            worksheet = workbook.copy_worksheet(template_sheet)
            worksheet.title = date_label
            add_worksheet_images(worksheet, template_images)

            entry = entry_by_date.get(date_key)
            if entry:
                fill_jb_contact_block_data_in_sheet(worksheet, entry, date_label, fab)

        if len(workbook.worksheets) > 1:
            workbook.remove(template_sheet)

        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)

        month_name = datetime(year, month, 1).strftime("%B")
        filename = f"{safe_filename('JB_Contact_Block_Maintenance')}_{safe_filename(fab)}_{month_name}_{year}.xlsx"
        return output, filename
    except Exception as e:
        log_progress(f"Error generating JB contact block maintenance report: {str(e)}")
        raise

