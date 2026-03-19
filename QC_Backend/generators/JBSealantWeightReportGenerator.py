from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import io
from datetime import datetime
import calendar
from paths import get_template_key, download_from_s3

def fill_jb_data_in_sheet(worksheet, entries, date):
    """Fill JB sealant weight data for a specific date into a worksheet"""
    try:
        # Sort entries by shift
        shift_order = {'A': 0, 'B': 1, 'C': 2}
        sorted_entries = sorted(entries, key=lambda e: shift_order.get(e.get('shift', ''), 3))
        
        current_row = 5  # Start from row 5 as per template
        
        for entry in sorted_entries:
            shift = entry.get('shift', '')
            lines = entry.get('lines', {})
            
            # Set date and shift for the 6 rows that will be filled for this shift (3 rows per line)
            for i in range(6):
                worksheet[f'A{current_row + i}'] = date
                worksheet[f'B{current_row + i}'] = f"Shift {shift}"
            
            # Process Line 1 (first 3 rows)
            line1_data = lines.get('1', {})
            line1_number = line1_data.get('line', '1')
            
            # Get JB position data for Line 1
            line1_positive_jb = line1_data.get('positiveJB', {})
            line1_middle_jb = line1_data.get('middleJB', {})
            line1_negative_jb = line1_data.get('negativeJB', {})
            
            # Fill Line 1 - Positive JB row
            line1_pos_row = current_row
            worksheet[f'C{line1_pos_row}'] = line1_number
            worksheet[f'D{line1_pos_row}'] = line1_data.get('po', '')
            worksheet[f'E{line1_pos_row}'] = line1_data.get('jbSupplier', '')
            worksheet[f'F{line1_pos_row}'] = line1_data.get('sealantSupplier', '')
            worksheet[f'G{line1_pos_row}'] = line1_data.get('sealantExpiry', '')
            worksheet[f'H{line1_pos_row}'] = '+ Ve JB'
            worksheet[f'I{line1_pos_row}'] = line1_positive_jb.get('jbWeight', '')
            worksheet[f'J{line1_pos_row}'] = line1_positive_jb.get('jbWeightWithSealant', '')
            worksheet[f'K{line1_pos_row}'] = line1_positive_jb.get('netSealantWeight', '')
            
            # Color code net sealant weight for Line 1 Positive JB
            try:
                if line1_positive_jb.get('netSealantWeight'):
                    weight_val = float(line1_positive_jb['netSealantWeight'])
                    if 4 <= weight_val <= 8:
                        worksheet[f'K{line1_pos_row}'].fill = PatternFill(start_color='92D050', end_color='92D050', fill_type='solid')
                    else:
                        worksheet[f'K{line1_pos_row}'].fill = PatternFill(start_color='FF9999', end_color='FF9999', fill_type='solid')
            except (ValueError, TypeError):
                pass
            
            # Fill Line 1 - Middle JB row
            line1_mid_row = current_row + 1
            worksheet[f'C{line1_mid_row}'] = line1_number
            worksheet[f'D{line1_mid_row}'] = line1_data.get('po', '')
            worksheet[f'E{line1_mid_row}'] = line1_data.get('jbSupplier', '')
            worksheet[f'F{line1_mid_row}'] = line1_data.get('sealantSupplier', '')
            worksheet[f'G{line1_mid_row}'] = line1_data.get('sealantExpiry', '')
            worksheet[f'H{line1_mid_row}'] = 'Middle JB'
            worksheet[f'I{line1_mid_row}'] = line1_middle_jb.get('jbWeight', '')
            worksheet[f'J{line1_mid_row}'] = line1_middle_jb.get('jbWeightWithSealant', '')
            worksheet[f'K{line1_mid_row}'] = line1_middle_jb.get('netSealantWeight', '')
            
            # Color code net sealant weight for Line 1 Middle JB
            try:
                if line1_middle_jb.get('netSealantWeight'):
                    weight_val = float(line1_middle_jb['netSealantWeight'])
                    if 4 <= weight_val <= 8:
                        worksheet[f'K{line1_mid_row}'].fill = PatternFill(start_color='92D050', end_color='92D050', fill_type='solid')
                    else:
                        worksheet[f'K{line1_mid_row}'].fill = PatternFill(start_color='FF9999', end_color='FF9999', fill_type='solid')
            except (ValueError, TypeError):
                pass
            
            # Fill Line 1 - Negative JB row
            line1_neg_row = current_row + 2
            worksheet[f'C{line1_neg_row}'] = line1_number
            worksheet[f'D{line1_neg_row}'] = line1_data.get('po', '')
            worksheet[f'E{line1_neg_row}'] = line1_data.get('jbSupplier', '')
            worksheet[f'F{line1_neg_row}'] = line1_data.get('sealantSupplier', '')
            worksheet[f'G{line1_neg_row}'] = line1_data.get('sealantExpiry', '')
            worksheet[f'H{line1_neg_row}'] = '- Ve JB'
            worksheet[f'I{line1_neg_row}'] = line1_negative_jb.get('jbWeight', '')
            worksheet[f'J{line1_neg_row}'] = line1_negative_jb.get('jbWeightWithSealant', '')
            worksheet[f'K{line1_neg_row}'] = line1_negative_jb.get('netSealantWeight', '')
            
            # Color code net sealant weight for Line 1 Negative JB
            try:
                if line1_negative_jb.get('netSealantWeight'):
                    weight_val = float(line1_negative_jb['netSealantWeight'])
                    if 4 <= weight_val <= 8:
                        worksheet[f'K{line1_neg_row}'].fill = PatternFill(start_color='92D050', end_color='92D050', fill_type='solid')
                    else:
                        worksheet[f'K{line1_neg_row}'].fill = PatternFill(start_color='FF9999', end_color='FF9999', fill_type='solid')
            except (ValueError, TypeError):
                pass
            
            # Total Module Weight for Line 1 - place it in the first row of Line 1
            worksheet[f'L{line1_pos_row}'] = line1_data.get('totalModuleWeight', '')
            
            # Add remarks for Line 1
            worksheet[f'M{line1_pos_row}'] = line1_data.get('remarks', '')
            
            # Process Line 2 (next 3 rows)
            line2_data = lines.get('2', {})
            line2_number = line2_data.get('line', '2')
            
            # Get JB position data for Line 2
            line2_positive_jb = line2_data.get('positiveJB', {})
            line2_middle_jb = line2_data.get('middleJB', {})
            line2_negative_jb = line2_data.get('negativeJB', {})
            
            # Fill Line 2 - Positive JB row
            line2_pos_row = current_row + 3
            worksheet[f'C{line2_pos_row}'] = line2_number
            worksheet[f'D{line2_pos_row}'] = line2_data.get('po', '')
            worksheet[f'E{line2_pos_row}'] = line2_data.get('jbSupplier', '')
            worksheet[f'F{line2_pos_row}'] = line2_data.get('sealantSupplier', '')
            worksheet[f'G{line2_pos_row}'] = line2_data.get('sealantExpiry', '')
            worksheet[f'H{line2_pos_row}'] = '+ Ve JB'
            worksheet[f'I{line2_pos_row}'] = line2_positive_jb.get('jbWeight', '')
            worksheet[f'J{line2_pos_row}'] = line2_positive_jb.get('jbWeightWithSealant', '')
            worksheet[f'K{line2_pos_row}'] = line2_positive_jb.get('netSealantWeight', '')
            
            # Color code net sealant weight for Line 2 Positive JB
            try:
                if line2_positive_jb.get('netSealantWeight'):
                    weight_val = float(line2_positive_jb['netSealantWeight'])
                    if 4 <= weight_val <= 8:
                        worksheet[f'K{line2_pos_row}'].fill = PatternFill(start_color='92D050', end_color='92D050', fill_type='solid')
                    else:
                        worksheet[f'K{line2_pos_row}'].fill = PatternFill(start_color='FF9999', end_color='FF9999', fill_type='solid')
            except (ValueError, TypeError):
                pass
            
            # Fill Line 2 - Middle JB row
            line2_mid_row = current_row + 4
            worksheet[f'C{line2_mid_row}'] = line2_number
            worksheet[f'D{line2_mid_row}'] = line2_data.get('po', '')
            worksheet[f'E{line2_mid_row}'] = line2_data.get('jbSupplier', '')
            worksheet[f'F{line2_mid_row}'] = line2_data.get('sealantSupplier', '')
            worksheet[f'G{line2_mid_row}'] = line2_data.get('sealantExpiry', '')
            worksheet[f'H{line2_mid_row}'] = 'Middle JB'
            worksheet[f'I{line2_mid_row}'] = line2_middle_jb.get('jbWeight', '')
            worksheet[f'J{line2_mid_row}'] = line2_middle_jb.get('jbWeightWithSealant', '')
            worksheet[f'K{line2_mid_row}'] = line2_middle_jb.get('netSealantWeight', '')
            
            # Color code net sealant weight for Line 2 Middle JB
            try:
                if line2_middle_jb.get('netSealantWeight'):
                    weight_val = float(line2_middle_jb['netSealantWeight'])
                    if 4 <= weight_val <= 8:
                        worksheet[f'K{line2_mid_row}'].fill = PatternFill(start_color='92D050', end_color='92D050', fill_type='solid')
                    else:
                        worksheet[f'K{line2_mid_row}'].fill = PatternFill(start_color='FF9999', end_color='FF9999', fill_type='solid')
            except (ValueError, TypeError):
                pass
            
            # Fill Line 2 - Negative JB row
            line2_neg_row = current_row + 5
            worksheet[f'C{line2_neg_row}'] = line2_number
            worksheet[f'D{line2_neg_row}'] = line2_data.get('po', '')
            worksheet[f'E{line2_neg_row}'] = line2_data.get('jbSupplier', '')
            worksheet[f'F{line2_neg_row}'] = line2_data.get('sealantSupplier', '')
            worksheet[f'G{line2_neg_row}'] = line2_data.get('sealantExpiry', '')
            worksheet[f'H{line2_neg_row}'] = '- Ve JB'
            worksheet[f'I{line2_neg_row}'] = line2_negative_jb.get('jbWeight', '')
            worksheet[f'J{line2_neg_row}'] = line2_negative_jb.get('jbWeightWithSealant', '')
            worksheet[f'K{line2_neg_row}'] = line2_negative_jb.get('netSealantWeight', '')
            
            # Color code net sealant weight for Line 2 Negative JB
            try:
                if line2_negative_jb.get('netSealantWeight'):
                    weight_val = float(line2_negative_jb['netSealantWeight'])
                    if 4 <= weight_val <= 8:
                        worksheet[f'K{line2_neg_row}'].fill = PatternFill(start_color='92D050', end_color='92D050', fill_type='solid')
                    else:
                        worksheet[f'K{line2_neg_row}'].fill = PatternFill(start_color='FF9999', end_color='FF9999', fill_type='solid')
            except (ValueError, TypeError):
                pass
            
            # Total Module Weight for Line 2 - place it in the first row of Line 2
            worksheet[f'L{line2_pos_row}'] = line2_data.get('totalModuleWeight', '')
            
            # Add remarks for Line 2
            worksheet[f'M{line2_pos_row}'] = line2_data.get('remarks', '')
            
            # Move to next shift (6 rows down - 3 rows per line)
            current_row += 6
        
        # Fill signatures at the bottom
        if sorted_entries and sorted_entries[0].get('signatures'):
            signatures = sorted_entries[0]['signatures']
            if signatures.get('preparedBy'):
                worksheet['C23'] = signatures.get('preparedBy', '')
            if signatures.get('verifiedBy'):
                worksheet['J23'] = signatures.get('verifiedBy', '')
        
        print(f"Filled JB sealant data for date {date}")
        
    except Exception as e:
        print(f"Error filling JB sealant data in sheet: {str(e)}")
        raise


def generate_jb_sealant_report(jb_data):
    """Generate JB sealant weight report with multiple sheets - one sheet per day of the month"""
    try:
        if not jb_data:
            raise ValueError("No JB sealant data provided")
        
        print("Received JB sealant data for report generation")
        
        # Handle different input formats
        if isinstance(jb_data, list):
            entries = jb_data
            year = datetime.now().year
            month = datetime.now().month
        else:
            entries = jb_data.get('entries', [])
            year = jb_data.get('year', datetime.now().year)
            month = jb_data.get('month', datetime.now().month)
        
        print(f"Entries count: {len(entries)}")
        print(f"Year: {year}, Month: {month}")
        
        # Group entries by date
        entries_by_date = {}
        for entry in entries:
            date = entry.get('date', '')
            if date not in entries_by_date:
                entries_by_date[date] = []
            entries_by_date[date].append(entry)
        
        # Get days in month
        days_in_month = calendar.monthrange(year, month)[1]
        
        # Load template
        template_key = get_template_key('Blank JB Sealant Weight Report.xlsx')
        template_path = download_from_s3(template_key)
        
        # Load workbook
        wb = load_workbook(template_path)
        
        # Create sheets for each day of the month
        for day in range(1, days_in_month + 1):
            date_str = f"{year}-{month:02d}-{day:02d}"
            formatted_date = f"{day:02d}.{month:02d}.{year}"
            
            # Create new sheet by copying template
            new_sheet = wb.copy_worksheet(wb.active)
            new_sheet.title = formatted_date
            
            # Get entries for this date
            date_entries = entries_by_date.get(date_str, [])
            
            if date_entries:
                # Fill data if entries exist
                fill_jb_data_in_sheet(new_sheet, date_entries, formatted_date)
        
        # Remove the original template sheet (first sheet)
        if len(wb.worksheets) > 1:
            wb.remove(wb.worksheets[0])
        
        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Generate filename
        month_name = datetime(year, month, 1).strftime("%B")
        filename = f"JB_Sealant_Weight_{month_name}_{year}.xlsx"
        
        print(f"JB sealant report generated successfully: {filename}")
        return output, filename
        
    except Exception as e:
        print(f"Error generating JB sealant report: {str(e)}")
        raise