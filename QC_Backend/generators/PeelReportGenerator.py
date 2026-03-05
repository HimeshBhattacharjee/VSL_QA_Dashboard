from openpyxl import load_workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side, NamedStyle)
import io
from paths import get_template_key, download_from_s3

def setup_peel_cell_styles(workbook):
    data_style = NamedStyle(name="peel_data_style")
    data_style.font = Font(name='Calibri', size=9)
    data_style.alignment = Alignment(horizontal='center', vertical='center')
    data_style.border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    header_style = NamedStyle(name="peel_header_style")
    header_style.font = Font(name='Calibri', size=9, bold=True)
    header_style.fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
    header_style.alignment = Alignment(horizontal='center', vertical='center')
    header_style.border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    highlight_style = NamedStyle(name="peel_highlight_style")
    highlight_style.font = Font(name='Calibri', size=9, color='FF0000')
    highlight_style.fill = PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid')
    highlight_style.alignment = Alignment(horizontal='center', vertical='center')
    highlight_style.border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    for style in [data_style, header_style, highlight_style]:
        if style.name not in workbook.named_styles:
            workbook.add_named_style(style)

def fill_peel_basic_info(worksheet, peel_data):
    try:
        form_data = peel_data.get('form_data', {})
        basic_info_mapping = {
            'preparedBySignature': 'E823',
            'verifiedBySignature': 'N823',
        }
        for field, cell_ref in basic_info_mapping.items():
            if field in form_data and form_data[field]:
                worksheet[cell_ref] = form_data[field]
                apply_peel_cell_formatting(worksheet[cell_ref], font_size=10, horizontal='center')
        print("Basic peel test information filled successfully")
    except Exception as e:
        print(f"Error filling basic peel test info: {str(e)}")
        raise

def fill_peel_test_data(worksheet, peel_data):
    try:
        form_data = peel_data.get('form_data', {})
        test_data_mapping = {}
        for rep in range(24):
            base_row = 7 + (rep * 34)
            test_data_mapping[f'row_{rep}_cell_0'] = f'B{base_row}'
            test_data_mapping[f'row_{rep}_cell_1'] = f'C{base_row}'
            test_data_mapping[f'row_{rep}_cell_2'] = f'D{base_row}'
            test_data_mapping[f'row_{rep}_cell_3'] = f'E{base_row}'
            test_data_mapping[f'row_{rep}_cell_4'] = f'F{base_row}'
            test_data_mapping[f'row_{rep}_cell_5'] = f'G{base_row}'
            for position in range(1, 17):
                row_offset = position
                for ribbon in range(1, 8):
                    cell_index = 6 + (position - 1) * 7 + (ribbon - 1)
                    col_offset = 8 + ribbon
                    test_data_mapping[f'row_{rep}_cell_{cell_index}'] = f'{get_column_letter(col_offset)}{base_row + row_offset}'
            for position in range(1, 17):
                row_offset = 17 + position
                for ribbon in range(1, 8):
                    cell_index = 118 + (position - 1) * 7 + (ribbon - 1)
                    col_offset = 8 + ribbon
                    test_data_mapping[f'row_{rep}_cell_{cell_index}'] = f'{get_column_letter(col_offset)}{base_row + row_offset}'
            for position in range(1, 17):
                row_offset = position
                test_data_mapping[f'front_avg_{rep}_{position}'] = f'P{base_row + row_offset}'
            for position in range(1, 17):
                row_offset = 17 + position
                test_data_mapping[f'back_avg_{rep}_{position}'] = f'P{base_row + row_offset}'
        for field, cell_ref in test_data_mapping.items():
            if field in form_data and form_data[field]:
                worksheet[cell_ref] = form_data[field]
                try:
                    value = float(form_data[field])
                    if value < 1.0:
                        apply_peel_cell_formatting(worksheet[cell_ref], font_size=9, horizontal='center', 
                                                 text_color='FF0000', fill_color='FFCCCC')
                    else:
                        apply_peel_cell_formatting(worksheet[cell_ref], font_size=9, horizontal='center')
                except (ValueError, TypeError):
                    apply_peel_cell_formatting(worksheet[cell_ref], font_size=9, horizontal='center')
        print("Peel test data filled successfully")
        
    except Exception as e:
        print(f"Error filling peel test data: {str(e)}")
        raise

def apply_peel_cell_formatting(cell, font_size=9, bold=False, 
                             text_color='000000', fill_color=None, 
                             horizontal='center', vertical='center'):
    cell.font = Font(name='Calibri', size=font_size, bold=bold, color=text_color)
    if fill_color:
        cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
    cell.alignment = Alignment(horizontal=horizontal, vertical=vertical, wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    cell.border = thin_border

def get_column_letter(col_idx):
    letters = []
    while col_idx > 0:
        col_idx, remainder = divmod(col_idx - 1, 26)
        letters.append(chr(65 + remainder))
    return ''.join(reversed(letters))

def generate_peel_filename(peel_data):
    report_name = peel_data.get('report_name', 'Peel_Test_Report')
    timestamp = peel_data.get('timestamp', '').split('T')[0] if peel_data.get('timestamp') else ''
    clean_name = "".join(c for c in report_name if c.isalnum() or c in (' ', '-', '_')).rstrip()
    clean_name = clean_name.replace(' ', '_')
    if timestamp:
        return f"Peel_Test_{clean_name}_{timestamp}.xlsx"
    else:
        return f"Peel_Test_{clean_name}.xlsx"

def generate_peel_report(peel_data):
    try:
        if not peel_data:
            raise ValueError("No peel test data provided")
        print("Received peel test data for report generation")
        print(f"Report name: {peel_data.get('report_name', 'N/A')}")
        print(f"Form data keys: {list(peel_data.get('form_data', {}).keys())}")
        template_key = get_template_key('Blank Solar Cell Peel Strength Test Report.xlsx')
        template_path = download_from_s3(template_key)
        wb = load_workbook(template_path)
        ws = wb.active
        setup_peel_cell_styles(wb)
        fill_peel_basic_info(ws, peel_data)
        fill_peel_test_data(ws, peel_data)
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        filename = generate_peel_filename(peel_data)
        print(f"Peel test report generated successfully: {filename}")
        return output, filename
    except Exception as e:
        print(f"Error generating peel test report: {str(e)}")
        raise