from logging_utils import log_progress
import calendar
from datetime import datetime
from functools import lru_cache
import io
import os
import re

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill

from generators.excel_image_utils import add_worksheet_images, collect_worksheet_images
from paths import get_template_key
from s3_service import get_s3_client
from services.shift_prepared_by_service import format_prepared_by


TEMPLATE_FILENAME = "Blank Peel Strength of Bus Ribbon to JB Soldering Test Report.xlsx"
FAB_LINES = {
    "FAB-II Line-I": ("Line - 1", "Line - 2"),
    "FAB-II Line-II": ("Line - 3", "Line - 4"),
}
SHIFT_ROW_START = {
    "A": 6,
    "B": 8,
    "C": 10,
}
READING_FIELDS = ("plusVe1", "plusVe2", "middle1", "middle2", "minusVe1", "minusVe2")
READING_COLUMNS = ("I", "J", "K", "L", "M", "N")
MIN_AVERAGE_N = 25
PASS_FILL = PatternFill(start_color="E2F0D9", end_color="E2F0D9", fill_type="solid")
FAIL_FILL = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
FAIL_FONT = Font(color="C00000")
NUMERIC_ONLY_RE = re.compile(r"^[+-]?(?:\d+(?:\.\d*)?|\.\d+)$")


def normalize_fab(fab):
    return "FAB-II Line-II" if fab == "FAB-II Line-II" else "FAB-II Line-I"


def safe_filename(value):
    clean_name = "".join(c for c in str(value or "") if c.isalnum() or c in (" ", "-", "_")).strip()
    return clean_name.replace(" ", "_") or "Peel_Strength_Bus_Ribbon_JB_Soldering"


def parse_numeric(value):
    if value in (None, ""):
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip()
    if not text or not NUMERIC_ONLY_RE.fullmatch(text):
        return None
    try:
        return float(text)
    except (TypeError, ValueError):
        return None


def to_excel_number(value):
    numeric_value = parse_numeric(value)
    if numeric_value is None:
        return ""
    return int(numeric_value) if numeric_value.is_integer() else numeric_value


def calculate_average(line_data):
    numeric_values = [value for value in (parse_numeric(line_data.get(field)) for field in READING_FIELDS) if value is not None]
    if not numeric_values:
        return None
    return round(sum(numeric_values) / len(numeric_values), 2)


def get_line_average(line_data):
    average = parse_numeric(line_data.get("average"))
    if average is not None:
        return round(average, 2)
    return calculate_average(line_data)


def get_line_remarks(line_data):
    average = get_line_average(line_data)
    if average is None:
        return ""
    return "OK" if average >= MIN_AVERAGE_N else "NOT OK"


@lru_cache(maxsize=1)
def get_template_bytes():
    s3_client = get_s3_client()
    bucket = os.getenv("S3_BUCKET_NAME")
    if not bucket:
        raise ValueError("S3_BUCKET_NAME environment variable not set")

    response = s3_client.get_object(Bucket=bucket, Key=get_template_key(TEMPLATE_FILENAME))
    return response["Body"].read()


def load_template_workbook():
    return load_workbook(io.BytesIO(get_template_bytes()))


def format_shift_signatures(entries, field):
    values_by_shift = []
    for entry in sorted(entries, key=lambda item: {"A": 0, "B": 1, "C": 2}.get(item.get("shift", ""), 9)):
        signatures = entry.get("signatures") or {}
        if field == "verifiedBy":
            value = signatures.get("verifiedBy") or signatures.get("reviewedBy")
        else:
            value = signatures.get(field)
        if value:
            values_by_shift.append((entry.get("shift", ""), value))

    if not values_by_shift:
        return ""

    unique_values = {value for _, value in values_by_shift}
    if len(unique_values) == 1:
        return values_by_shift[0][1]
    return "; ".join(f"{shift}: {value}" for shift, value in values_by_shift if shift)


def apply_average_style(cell, average):
    if average is None:
        return
    if average >= MIN_AVERAGE_N:
        cell.fill = PASS_FILL
    else:
        cell.fill = FAIL_FILL
        cell.font = FAIL_FONT


def write_line_row(worksheet, row, date_label, shift, line_label, line_data):
    worksheet[f"B{row}"] = date_label
    worksheet[f"C{row}"] = shift
    worksheet[f"D{row}"] = line_label[len(line_label) - 1]
    if str(line_data.get("status", "ON")).upper() == "OFF":
        for column in "EFGHIJKLMNOP":
            worksheet[f"{column}{row}"] = "OFF"
        return
    worksheet[f"E{row}"] = line_data.get("po", "")
    worksheet[f"F{row}"] = line_data.get("jbStatus", "")
    worksheet[f"G{row}"] = line_data.get("busRibbonStatus", "")
    worksheet[f"H{row}"] = line_data.get("busRibbonDimension", "")

    for field, column in zip(READING_FIELDS, READING_COLUMNS):
        worksheet[f"{column}{row}"] = to_excel_number(line_data.get(field))

    average = get_line_average(line_data)
    average_cell = worksheet[f"O{row}"]
    average_cell.value = "" if average is None else to_excel_number(average)
    apply_average_style(average_cell, average)
    worksheet[f"P{row}"] = get_line_remarks(line_data)


def fill_peel_strength_data_in_sheet(worksheet, entries, date_label, fab):
    entry_by_shift = {entry.get("shift"): entry for entry in entries if entry.get("shift") in SHIFT_ROW_START}
    line_labels = FAB_LINES[normalize_fab(fab)]

    for shift, start_row in SHIFT_ROW_START.items():
        entry = entry_by_shift.get(shift)
        if not entry:
            continue

        lines = entry.get("lines") or {}
        for offset, line_label in enumerate(line_labels):
            write_line_row(worksheet, start_row + offset, date_label, shift, line_label, lines.get(line_label, {}) or {})

    prepared_by = format_prepared_by(entries)
    verified_by = format_shift_signatures(entries, "verifiedBy")
    if prepared_by:
        worksheet["D12"] = prepared_by
    if verified_by:
        worksheet["L12"] = verified_by


def generate_peel_strength_bus_ribbon_jb_soldering_report(report_data):
    try:
        if not report_data:
            raise ValueError("No peel strength bus ribbon JB soldering data provided")

        if isinstance(report_data, list):
            entries = report_data
            year = datetime.now().year
            month = datetime.now().month
            fab = "FAB-II Line-I"
        else:
            entries = report_data.get("entries", [])
            year = int(report_data.get("year", datetime.now().year))
            month = int(report_data.get("month", datetime.now().month))
            fab = normalize_fab(report_data.get("fab"))

        entries_by_date = {}
        for entry in entries:
            if normalize_fab(entry.get("fab")) != fab:
                continue
            date = str(entry.get("date", "")).split("T")[0]
            if date:
                entries_by_date.setdefault(date, []).append(entry)

        workbook = load_template_workbook()
        template_sheet = workbook.active
        template_images = collect_worksheet_images(template_sheet)

        days_in_month = calendar.monthrange(year, month)[1]
        for day in range(1, days_in_month + 1):
            date_key = f"{year}-{month:02d}-{day:02d}"
            date_label = f"{day:02d}.{month:02d}.{year}"
            worksheet = workbook.copy_worksheet(template_sheet)
            worksheet.title = date_label
            add_worksheet_images(worksheet, template_images)

            date_entries = entries_by_date.get(date_key, [])
            if date_entries:
                fill_peel_strength_data_in_sheet(worksheet, date_entries, date_label, fab)

        if len(workbook.worksheets) > 1:
            workbook.remove(template_sheet)

        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)

        month_name = datetime(year, month, 1).strftime("%B")
        filename = (
            f"{safe_filename('Peel_Strength_Bus_Ribbon_JB_Soldering')}_"
            f"{safe_filename(fab)}_{month_name}_{year}.xlsx"
        )
        return output, filename
    except Exception as e:
        log_progress(f"Error generating peel strength bus ribbon JB soldering report: {str(e)}")
        raise

