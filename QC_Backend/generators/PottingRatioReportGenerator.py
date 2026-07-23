from logging_utils import log_progress
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import io
from datetime import datetime
import calendar
from paths import get_template_key, download_from_s3
from generators.excel_image_utils import add_worksheet_images, collect_worksheet_images
from services.potting_ratio_evaluation import evaluate_potting_ratio

SUCCESS_FILL = PatternFill(start_color='92D050', end_color='92D050', fill_type='solid')
FAILURE_FILL = PatternFill(start_color='FF9999', end_color='FF9999', fill_type='solid')
NO_FILL = PatternFill(fill_type=None)

def get_display_line_numbers(line_group):
    return ('3', '4') if line_group == 'Line-II' else ('1', '2')

def fill_potting_data_in_sheet(worksheet, entries, date):
    """Fill potting ratio data for a specific date into a worksheet"""
    try:
        # Sort entries by shift
        shift_order = {'A': 0, 'B': 1, 'C': 2}
        sorted_entries = sorted(entries, key=lambda e: shift_order.get(e.get('shift', ''), 3))
        
        current_row = 6  # Start from row 6 as per template
        
        for entry in sorted_entries:
            shift = entry.get('shift', '')
            lines = entry.get('lines', {})
            display_line_1, display_line_2 = get_display_line_numbers(entry.get('lineGroup'))
            entry_signatures = entry.get('signatures', {})
            
            # Set date in B column
            worksheet[f'B{current_row}'] = date
            worksheet[f'B{current_row+1}'] = date
            
            # Set shift in C column
            worksheet[f'C{current_row}'] = f"Shift {shift}"
            worksheet[f'C{current_row+1}'] = f"Shift {shift}"
            
            # Line 1 (first row)
            line1 = lines.get('1', {})
            worksheet[f'D{current_row}'] = display_line_1
            worksheet[f'E{current_row}'] = line1.get('po', '')  # PO
            worksheet[f'F{current_row}'] = line1.get('pottingSupplier', '')
            worksheet[f'G{current_row}'] = line1.get('partA', '')
            worksheet[f'H{current_row}'] = line1.get('partB', '')
            worksheet[f'I{current_row}'] = line1.get('ratio', '') and f"{line1.get('ratio', '')}:1" or ''
            worksheet[f'I{current_row}'].fill = NO_FILL
            
            worksheet[f'J{current_row}'] = line1.get('totalWeight', '')
            line1_evaluation = evaluate_potting_ratio(line1.get('ratio'))
            worksheet[f'K{current_row}'] = line1_evaluation.status
            worksheet[f'K{current_row}'].fill = NO_FILL
            if line1_evaluation.status:
                worksheet[f'K{current_row}'].fill = SUCCESS_FILL if line1_evaluation.status == 'OK' else FAILURE_FILL
            if str(line1.get('status', 'ON')).upper() == 'OFF':
                for column in 'EFGHIJK':
                    worksheet[f'{column}{current_row}'] = 'OFF'
            
            # Line 2 (second row)
            line2 = lines.get('2', {})
            worksheet[f'D{current_row+1}'] = display_line_2
            worksheet[f'E{current_row+1}'] = line2.get('po', '')  # PO
            worksheet[f'F{current_row+1}'] = line2.get('pottingSupplier', '')
            worksheet[f'G{current_row+1}'] = line2.get('partA', '')
            worksheet[f'H{current_row+1}'] = line2.get('partB', '')
            worksheet[f'I{current_row+1}'] = line2.get('ratio', '') and f"{line2.get('ratio', '')}:1" or ''
            worksheet[f'I{current_row+1}'].fill = NO_FILL
            
            worksheet[f'J{current_row+1}'] = line2.get('totalWeight', '')
            line2_evaluation = evaluate_potting_ratio(line2.get('ratio'))
            worksheet[f'K{current_row+1}'] = line2_evaluation.status
            worksheet[f'K{current_row+1}'].fill = NO_FILL
            if line2_evaluation.status:
                worksheet[f'K{current_row+1}'].fill = SUCCESS_FILL if line2_evaluation.status == 'OK' else FAILURE_FILL
            if str(line2.get('status', 'ON')).upper() == 'OFF':
                for column in 'EFGHIJK':
                    worksheet[f'{column}{current_row+1}'] = 'OFF'
            
            # Fill signatures at the bottom (one set per shift)
            if entry_signatures.get('preparedBy'):
                worksheet['D12'] = entry_signatures.get('preparedBy', '')
            if entry_signatures.get('verifiedBy'):
                worksheet['J12'] = entry_signatures.get('verifiedBy', '')
            
            # Move to next shift (2 rows down)
            current_row += 2
        
        log_progress(f"Filled potting data for date {date}")
        
    except Exception as e:
        log_progress(f"Error filling potting data in sheet: {str(e)}")
        raise


def generate_potting_report(potting_data):
    """Generate potting ratio report with multiple sheets - one sheet per day of the month"""
    try:
        if not potting_data:
            raise ValueError("No potting data provided")
        
        log_progress("Received potting data for report generation")
        
        # Handle different input formats
        if isinstance(potting_data, list):
            entries = potting_data
            year = datetime.now().year
            month = datetime.now().month
        else:
            entries = potting_data.get('entries', [])
            year = potting_data.get('year', datetime.now().year)
            month = potting_data.get('month', datetime.now().month)
        
        log_progress(f"Entries count: {len(entries)}")
        log_progress(f"Year: {year}, Month: {month}")
        
        # Group entries by date
        entries_by_date = {}
        for entry in entries:
            date = entry.get('date', '')
            if date not in entries_by_date:
                entries_by_date[date] = []
            entries_by_date[date].append(entry)
        
        # Get days in month
        days_in_month = calendar.monthrange(year, month)[1]
        
        # Load template
        template_key = get_template_key('Blank Potting Ratio Measurement Report.xlsx')
        template_path = download_from_s3(template_key)
        
        # Load workbook
        wb = load_workbook(template_path)
        
        # Get the template sheet
        template_sheet = wb.active
        template_images = collect_worksheet_images(template_sheet)
        
        # Create sheets for each day of the month
        for day in range(1, days_in_month + 1):
            date_str = f"{year}-{month:02d}-{day:02d}"
            formatted_date = f"{day:02d}.{month:02d}.{year}"
            
            # Create new sheet by copying template
            new_sheet = wb.copy_worksheet(template_sheet)
            new_sheet.title = formatted_date
            add_worksheet_images(new_sheet, template_images)
            
            # Get entries for this date
            date_entries = entries_by_date.get(date_str, [])
            
            if date_entries:
                # Fill data if entries exist
                fill_potting_data_in_sheet(new_sheet, date_entries, formatted_date)
        
        # Remove the original template sheet
        wb.remove(template_sheet)
        
        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Generate filename
        month_name = datetime(year, month, 1).strftime("%B")
        filename = f"Potting_Ratio_Measurement_{month_name}_{year}.xlsx"
        
        log_progress(f"Potting report generated successfully: {filename}")
        return output, filename
        
    except Exception as e:
        log_progress(f"Error generating potting report: {str(e)}")
        raise

