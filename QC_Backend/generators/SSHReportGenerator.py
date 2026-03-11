from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import io
from datetime import datetime
import re
from paths import get_template_key, download_from_s3

def fill_ssh_test_data(worksheet, entries):
    try:
        def sort_key(entry):
            date = entry.get('date', '')
            shift = entry.get('shift', '')
            shift_order = {'A': 0, 'B': 1, 'C': 2}
            return (date, shift_order.get(shift, 3))
        
        sorted_entries = sorted(entries, key=sort_key)
        entries_by_date = {}
        
        # Group entries by date
        for entry in sorted_entries:
            date = entry.get('date', '')
            if date not in entries_by_date:
                entries_by_date[date] = {}
            shift = entry.get('shift', '')
            entries_by_date[date][shift] = entry
            
        current_row = 6
        print(f"Filling test data for {len(entries_by_date)} dates starting at row {current_row}")
        
        # Process each date
        for date, shift_entries in entries_by_date.items():
            # Format date for display (convert from YYYY-MM-DD to DD.MM.YYYY)
            try:
                date_obj = datetime.strptime(date, '%Y-%m-%d')
                formatted_date = date_obj.strftime("%d.%m.%Y")
            except:
                formatted_date = date
            
            # Set date in C column - will be merged from row to row+5 in template
            worksheet[f'C{current_row}'] = formatted_date
            
            # Process each shift in order (A, B, C)
            all_shifts = ['A', 'B', 'C']
            
            for shift_index, shift in enumerate(all_shifts):
                # Calculate rows for this shift
                # Shift A: rows current_row and current_row+1
                # Shift B: rows current_row+2 and current_row+3
                # Shift C: rows current_row+4 and current_row+5
                shift_start_row = current_row + (shift_index * 2)
                shift_end_row = shift_start_row + 1
                
                # Set shift in column D (only in first row of shift as template has merged cells)
                worksheet[f'D{shift_start_row}'] = f"Shift {shift}"
                
                # Get entry for this shift if exists
                shift_entry = shift_entries.get(shift)
                
                if shift_entry and shift_entry.get('lines'):
                    # Set PO number (same for both lines of this shift)
                    po_value = shift_entry.get('po', '')
                    worksheet[f'E{shift_start_row}'] = po_value
                    
                    # Get checkedBy from entry
                    checked_by = shift_entry.get('checkedBy', '')
                    
                    # Line 1 (first row of shift)
                    line1 = shift_entry['lines'].get('1', {})
                    fill_line_row(worksheet, shift_start_row, line1, checked_by)
                    
                    # Line 2 (second row of shift)
                    line2 = shift_entry['lines'].get('2', {})
                    fill_line_row(worksheet, shift_end_row, line2, checked_by)
                else:
                    # Fill empty rows for missing shift
                    worksheet[f'E{shift_start_row}'] = ''  # Empty PO for missing shift
                    fill_line_row(worksheet, shift_start_row, {}, '')
                    fill_line_row(worksheet, shift_end_row, {}, '')
            
            # Move to next date (6 rows down as per template: 3 shifts × 2 rows each)
            current_row += 6
        
        print(f"Filled test data successfully up to row {current_row - 1}")
        
    except Exception as e:
        print(f"Error filling SSH test data: {str(e)}")
        raise

def fill_line_row(worksheet, row, line_data, checked_by=''):
    """Fill a single row with line data"""
    try:        
        # Line Number (Column F) - this should be the line number from the data
        worksheet[f'F{row}'] = line_data.get('line', '')

        # Sealant Supplier (Column G)
        worksheet[f'G{row}'] = line_data.get('sealantSupplier', '')
        
        # Sealant Expiry Date (Column H)
        exp_date = line_data.get('sealantExpDate', '')
        if exp_date:
            try:
                # Try to format date if it's in YYYY-MM-DD format
                if re.match(r'\d{4}-\d{2}-\d{2}', exp_date):
                    date_obj = datetime.strptime(exp_date, '%Y-%m-%d')
                    worksheet[f'H{row}'] = date_obj.strftime("%d.%m.%Y")
                else:
                    worksheet[f'H{row}'] = exp_date
            except:
                worksheet[f'H{row}'] = exp_date
        
        # Sample Taking Time (Column I)
        worksheet[f'I{row}'] = line_data.get('sampleTakingTime', '')
        
        # Sample Testing Time (Column J)
        worksheet[f'J{row}'] = line_data.get('sampleTestingTime', '')
        
        # Test Result with color coding (Column K)
        result = line_data.get('result', '')
        result_cell = worksheet[f'K{row}']
        result_cell.value = result
        
        # Apply color coding based on result
        try:
            # Try to convert to float for numeric comparison
            if result and str(result).replace('.', '').isdigit():
                result_float = float(result)
                if result_float >= 39:
                    result_cell.fill = PatternFill(start_color='92D050', end_color='92D050', fill_type='solid')
                elif result_float < 39 and result_float > 0:
                    result_cell.fill = PatternFill(start_color='FF9999', end_color='FF9999', fill_type='solid')
            else:
                # Handle text results like "Pass"/"Fail"
                if result and str(result).lower() == 'pass':
                    result_cell.fill = PatternFill(start_color='92D050', end_color='92D050', fill_type='solid')
                elif result and str(result).lower() == 'fail':
                    result_cell.fill = PatternFill(start_color='FF9999', end_color='FF9999', fill_type='solid')
        except (ValueError, TypeError):
            pass
        
        # Checked By (Column L)
        worksheet[f'L{row}'] = line_data.get('checkedBy', checked_by)
        
        # Remarks (Column M)
        worksheet[f'M{row}'] = line_data.get('remarks', '')
            
    except Exception as e:
        print(f"Error filling line row at row {row}: {str(e)}")
        raise

def fill_ssh_signatures(worksheet, form_data):
    """Fill signature data into the worksheet"""
    try:
        # Fill Prepared By at row 193 (as per template)
        prepared_by = form_data.get('preparedBy', '') or form_data.get('preparedBySignature', '')
        if prepared_by:
            # According to template, Prepared By is at row 193, column E
            cell = worksheet['E193']
            cell.value = prepared_by
        
        # Fill Approved By at row 193 (as per template)
        approved_by = form_data.get('approvedBy', '') or form_data.get('approvedBySignature', '')
        if approved_by:
            # According to template, Approved By is at row 193, column K
            cell = worksheet['K193']
            cell.value = approved_by
        
        print("SSH signatures filled successfully")
        
    except Exception as e:
        print(f"Error filling SSH signatures: {str(e)}")
        raise

def generate_ssh_report(ssh_data):
    """Generate SSH test report from MongoDB data structure"""
    try:
        if not ssh_data:
            raise ValueError("No SSH test data provided")
        
        print("Received SSH test data for report generation")
        
        # Handle different input formats
        if isinstance(ssh_data, list):
            entries = ssh_data
            form_data = {}
            report_name = "SSH_Test_Report"
        else:
            entries = ssh_data.get('entries', [])
            form_data = ssh_data.get('form_data', {})
            report_name = ssh_data.get('name', ssh_data.get('report_name', 'SSH_Test_Report'))
        
        # If entries is empty but ssh_data itself might be an entry
        if not entries and isinstance(ssh_data, dict) and 'shift' in ssh_data and 'lines' in ssh_data:
            entries = [ssh_data]
        
        print(f"Entries count: {len(entries)}")
        
        # Load template
        template_key = get_template_key('Blank Sealant Shore Hardness Test Report.xlsx')
        template_path = download_from_s3(template_key)
        
        # Load workbook
        wb = load_workbook(template_path)
        ws = wb.active
        
        # Fill data
        fill_ssh_test_data(ws, entries)
        
        # Fill signatures
        if form_data:
            fill_ssh_signatures(ws, form_data)
        
        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Generate filename
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        if len(entries) == 1:
            date_str = entries[0].get('date', '').replace('-', '')
            clean_name = f"SSH_Report_{date_str}"
        else:
            # Get date range if multiple entries
            dates = sorted(list(set(e.get('date', '') for e in entries if e.get('date'))))
            if dates:
                start_date = dates[0].replace('-', '')
                end_date = dates[-1].replace('-', '')
                clean_name = f"SSH_Report_{start_date}_to_{end_date}"
            else:
                clean_name = "SSH_Test_Report"
        
        filename = f"{clean_name}_{timestamp}.xlsx"
        
        print(f"SSH test report generated successfully: {filename}")
        return output, filename
        
    except Exception as e:
        print(f"Error generating SSH test report: {str(e)}")
        raise