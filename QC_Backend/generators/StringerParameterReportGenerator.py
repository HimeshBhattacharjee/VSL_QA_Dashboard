from copy import copy
from functools import lru_cache
import io
import os
import re

from openpyxl import load_workbook

from generators.stringer_parameter_mappings import (
    COMMON_EXCEL_COLUMNS,
    get_stringer_parameter_mapper,
)
from paths import get_template_key
from s3_service import get_s3_client


NUMERIC_PATTERN = re.compile(r"^[+-]?(?:\d+(?:\.\d+)?|\.\d+)$")
DATA_START_ROW = 5


@lru_cache(maxsize=2)
def get_template_bytes(line: str) -> bytes:
    mapper = get_stringer_parameter_mapper(line)
    bucket = os.getenv("S3_BUCKET_NAME")
    if not bucket:
        raise ValueError("S3_BUCKET_NAME environment variable not set")
    response = get_s3_client().get_object(
        Bucket=bucket,
        Key=get_template_key(mapper.template_filename),
    )
    return response["Body"].read()


def load_template_workbook(line: str):
    return load_workbook(io.BytesIO(get_template_bytes(line)))


def to_excel_value(value):
    if value in (None, ""):
        return ""
    if isinstance(value, bool) or isinstance(value, (int, float)):
        return value
    text = str(value).strip()
    if NUMERIC_PATTERN.fullmatch(text):
        numeric_value = float(text)
        return int(numeric_value) if numeric_value.is_integer() else numeric_value
    return text


def get_machine_pair_ranges(worksheet):
    return sorted(
        (
            merged_range
            for merged_range in worksheet.merged_cells.ranges
            if merged_range.min_col == 7
            and merged_range.max_col == 7
            and merged_range.min_row >= DATA_START_ROW
            and merged_range.max_row == merged_range.min_row + 1
        ),
        key=lambda merged_range: merged_range.min_row,
    )


def shift_merged_range(range_boundaries, row_offset: int, insert_at: int):
    min_col, min_row, max_col, max_row = range_boundaries
    if min_row >= insert_at:
        min_row += row_offset
        max_row += row_offset
    elif min_row < insert_at <= max_row:
        max_row += row_offset
    return min_col, min_row, max_col, max_row


def copy_cell_style(source, target):
    target._style = copy(source._style)
    target.font = copy(source.font)
    target.fill = copy(source.fill)
    target.border = copy(source.border)
    target.alignment = copy(source.alignment)
    target.number_format = source.number_format
    target.protection = copy(source.protection)


def extend_machine_pair_capacity(worksheet, pair_ranges, required_pairs: int):
    extra_pairs = required_pairs - len(pair_ranges)
    if extra_pairs <= 0:
        return

    template_pair = pair_ranges[-1]
    pair_height = template_pair.max_row - template_pair.min_row + 1
    insert_at = template_pair.max_row + 1
    extra_rows = extra_pairs * pair_height
    merged_ranges = [merged_range.bounds for merged_range in worksheet.merged_cells.ranges]
    template_pair_merges = [
        bounds
        for bounds in merged_ranges
        if bounds[1] >= template_pair.min_row and bounds[3] <= template_pair.max_row
    ]

    for merged_range in list(worksheet.merged_cells.ranges):
        worksheet.unmerge_cells(str(merged_range))

    worksheet.insert_rows(insert_at, extra_rows)

    for offset in range(extra_rows):
        source_row = template_pair.min_row + (offset % pair_height)
        target_row = insert_at + offset
        worksheet.row_dimensions[target_row].height = worksheet.row_dimensions[source_row].height
        for column_index in range(1, worksheet.max_column + 1):
            source = worksheet.cell(row=source_row, column=column_index)
            target = worksheet.cell(row=target_row, column=column_index)
            copy_cell_style(source, target)
            target.value = None

    for min_col, min_row, max_col, max_row in (
        shift_merged_range(bounds, extra_rows, insert_at)
        for bounds in merged_ranges
    ):
        worksheet.merge_cells(
            start_row=min_row,
            start_column=min_col,
            end_row=max_row,
            end_column=max_col,
        )

    for pair_index in range(extra_pairs):
        row_offset = insert_at - template_pair.min_row + (pair_index * pair_height)
        for min_col, min_row, max_col, max_row in template_pair_merges:
            worksheet.merge_cells(
                start_row=min_row + row_offset,
                start_column=min_col,
                end_row=max_row + row_offset,
                end_column=max_col,
            )


def ensure_machine_pair_capacity(worksheet, required_pairs: int):
    pair_ranges = get_machine_pair_ranges(worksheet)
    if not pair_ranges:
        raise ValueError("Stringer Parameter template has no machine row pairs")
    if required_pairs > len(pair_ranges):
        extend_machine_pair_capacity(worksheet, pair_ranges, required_pairs)
        pair_ranges = get_machine_pair_ranges(worksheet)
    if required_pairs > len(pair_ranges):
        raise ValueError(
            f"Stringer Parameter template supports {len(pair_ranges)} rows after expansion, "
            f"but {required_pairs} rows were requested"
        )
    return pair_ranges[:required_pairs]


def write_common_fields(worksheet, row: int, report_row: dict):
    date_value = str(report_row.get("date") or "")
    date_parts = date_value.split("-")
    formatted_date = (
        f"{date_parts[2]}/{date_parts[1]}/{date_parts[0]}"
        if len(date_parts) == 3
        else date_value
    )
    common_values = {
        "date": formatted_date,
        "shift": str(report_row.get("shift") or ""),
        "poNumber": str(report_row.get("poNumber") or ""),
        "moduleType": str(report_row.get("moduleType") or ""),
        "cellType": str(report_row.get("cellType") or ""),
        "cellWp": to_excel_value(report_row.get("cellWp", "")),
        "machine": to_excel_value(report_row.get("machine", "")),
    }
    for field, value in common_values.items():
        worksheet[f"{COMMON_EXCEL_COLUMNS[field]}{row}"] = value


def write_unit_values(worksheet, row: int, unit_label: str, values: dict, mapper):
    worksheet[f"{COMMON_EXCEL_COLUMNS['unit']}{row}"] = unit_label
    for column in mapper.audit_columns:
        worksheet[f"{column['excel']}{row}"] = to_excel_value(values.get(column["key"], ""))


def generate_stringer_parameter_report(report_data: dict):
    if not report_data:
        raise ValueError("No Stringer Parameter Report data provided")

    line = str(report_data.get("line") or "I")
    mapper = get_stringer_parameter_mapper(line)
    rows = report_data.get("rows") or []
    year = int(report_data.get("year"))
    month = int(report_data.get("month"))

    workbook = load_template_workbook(mapper.line)
    worksheet = workbook.active
    pair_ranges = ensure_machine_pair_capacity(worksheet, len(rows))

    for report_row, pair_range in zip(rows, pair_ranges):
        top_row = pair_range.min_row
        bottom_row = pair_range.max_row
        write_common_fields(worksheet, top_row, report_row)
        audit_values = report_row.get("auditValues") or {}
        write_unit_values(worksheet, top_row, "A", audit_values.get("unitA") or {}, mapper)
        write_unit_values(worksheet, bottom_row, "B", audit_values.get("unitB") or {}, mapper)

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    filename = f"Stringer_Parameter_Report_Line-{mapper.line}_{year}_{month:02d}.xlsx"
    return output, filename
