from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import io
from datetime import datetime
from paths import get_template_key, download_from_s3

def fill_wet_leakage_test_data(worksheet, entries):
    try:
        sorted_entries = sorted(entries, key=lambda x: x.get('testingDate', ''))
        start_row = 6
        max_rows = 31
        print(f"Filling {len(sorted_entries)} test data rows starting at row {start_row}")
        for idx, entry in enumerate(sorted_entries[:max_rows]):
            row = start_row + idx
            testing_date = entry.get('testingDate', '')
            if testing_date:
                try:
                    if 'T' in testing_date:
                        date_obj = datetime.fromisoformat(testing_date.replace('Z', '+00:00'))
                        worksheet[f'B{row}'] = date_obj.strftime("%d.%m.%Y")
                    else:
                        for fmt in ['%Y-%m-%d', '%d.%m.%Y', '%d/%m/%Y']:
                            try:
                                date_obj = datetime.strptime(testing_date, fmt)
                                worksheet[f'B{row}'] = date_obj.strftime("%d.%m.%Y")
                                break
                            except:
                                continue
                        else:
                            worksheet[f'B{row}'] = testing_date
                except:
                    worksheet[f'B{row}'] = testing_date
            worksheet[f'C{row}'] = entry.get('po', '')
            worksheet[f'D{row}'] = entry.get('moduleType', '')
            worksheet[f'F{row}'] = entry.get('moduleNo', '')
            worksheet[f'G{row}'] = entry.get('cellSupplier', '')
            worksheet[f'H{row}'] = entry.get('encapsulantSupplier', '')
            worksheet[f'I{row}'] = entry.get('rearGlassSupplier', '')
            worksheet[f'J{row}'] = entry.get('jbSupplier', '')
            worksheet[f'K{row}'] = entry.get('adhesiveSealantSupplier', '')
            worksheet[f'L{row}'] = entry.get('pottingSealantSupplier', '')
            worksheet[f'M{row}'] = entry.get('waterTemp', '')
            worksheet[f'N{row}'] = entry.get('waterResistivity', '')
            worksheet[f'O{row}'] = entry.get('IR', '')
            result = entry.get('result', '')
            worksheet[f'P{row}'] = result
            if result == 'Pass':
                worksheet[f'P{row}'].fill = PatternFill(start_color='92D050', end_color='92D050', fill_type='solid')
            elif result == 'Fail':
                worksheet[f'P{row}'].fill = PatternFill(start_color='FF9999', end_color='FF9999', fill_type='solid')
            worksheet[f'Q{row}'] = entry.get('testDoneBy', '')
        print(f"Filled {min(len(sorted_entries), max_rows)} test data rows successfully")
    except Exception as e:
        print(f"Error filling Wet Leakage test data: {str(e)}")
        raise

def fill_wet_leakage_signatures(worksheet, form_data):
    try:
        prepared_by = form_data.get('preparedBySignature', '')
        if prepared_by:
            worksheet['C37'] = prepared_by
        reviewed_by = form_data.get('reviewedBySignature', '')
        if reviewed_by:
            worksheet['I37'] = reviewed_by
        approved_by = form_data.get('approvedBySignature', '')
        if approved_by:
            worksheet['N37'] = approved_by
        print("Wet Leakage signatures filled successfully")
    except Exception as e:
        print(f"Error filling Wet Leakage signatures: {str(e)}")
        raise

def generate_wet_leakage_report(wet_leakage_data):
    try:
        if not wet_leakage_data:
            raise ValueError("No Wet Leakage test data provided")
        print("Received Wet Leakage test data for report generation")
        print(f"Report name: {wet_leakage_data.get('name', 'N/A')}")
        print(f"Entries count: {len(wet_leakage_data.get('entries', []))}")
        print(f"Form data keys: {list(wet_leakage_data.get('form_data', {}).keys())}")
        entries = wet_leakage_data.get('entries', [])
        form_data = wet_leakage_data.get('form_data', {})
        template_key = get_template_key('Blank Wet Leakage Test Report.xlsx')
        template_path = download_from_s3(template_key)
        wb = load_workbook(template_path)
        ws = wb.active
        fill_wet_leakage_test_data(ws, entries)
        fill_wet_leakage_signatures(ws, form_data)
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        report_name = wet_leakage_data.get('name', wet_leakage_data.get('report_name', 'Wet_Leakage_Test_Report'))
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        clean_name = "".join(c for c in report_name if c.isalnum() or c in (' ', '-', '_')).rstrip()
        clean_name = clean_name.replace(' ', '_')
        filename = f"{clean_name}_{timestamp}.xlsx"
        print(f"Wet Leakage test report generated successfully: {filename}")
        return output, filename
    except Exception as e:
        print(f"Error generating Wet Leakage test report: {str(e)}")
        raise