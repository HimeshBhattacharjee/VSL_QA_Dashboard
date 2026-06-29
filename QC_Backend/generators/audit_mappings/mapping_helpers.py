import re

from openpyxl.utils import get_column_letter

LINE_I_MACHINE_TEMP_FIELDS = (
    'Flux Temp',
    'Preheat base-1',
    'Preheat base-2',
    'Solder base-1',
    'Solder base-2',
    'Holding base-1',
    'Combined Plates',
    'Holding base-2',
    'Holding base-3',
    'Drying base-1',
    'Drying base-2',
    'Drying base-3',
    'Drying base-4',
    'Drying base-5',
)

LINE_I_LIGHT_INTENSITY_FIELDS = (
    'Solder Time ms',
    'Solder Temp ˚C',
    '#1',
    '#2',
    '#3',
    '#4',
    '#5',
    '#6',
    '#7',
    '#8',
    '#9',
    '#10',
    '#11',
    '#12',
    '#13',
    '#14',
    '#15',
    '#16',
    '#17',
    '#18',
    '#19',
    '#20',
)

LINE_I_STRINGER_SETUP_ROWS = (
    (1, 58, 59),
    (2, 60, 61),
    (3, 62, 63),
    (4, 64, 65),
    (5, 66, 67),
    (6, 68, 69),
)

LINE_I_LIGHT_INTENSITY_ROWS = (
    (1, 72, 73),
    (2, 74, 75),
    (3, 76, 77),
    (4, 78, 79),
    (5, 80, 81),
    (6, 82, 83),
)

LINE_I_AUTO_BUSSING_SOLDERING_TIME_ROWS = {
    'Line-1': {'front': 134, 'middle': 135, 'back': 136},
    'Line-2': {'front': 149, 'middle': 150, 'back': 151},
    'Line-3': {'front': 164, 'middle': 165, 'back': 166},
}

LINE_I_AUTO_BUSSING_SOLDERING_TIME_COLUMNS = ('I', 'M', 'Q', 'U', 'Y', 'AC')

AUTO_BUSSING_PATCH_PARAMETER_ID = '7-10'

LINE_I_AUTO_BUSSING_BLOCK_ROWS = {
    'Line-1': 133,
    'Line-2': 148,
    'Line-3': 163,
}

AUTO_BUSSING_PATCH_VALUE_COLUMNS = {
    'sample_1': {
        'length': 'I',
        'height': 'M',
        'width': 'Q',
    },
    'sample_2': {
        'length': 'U',
        'height': 'Y',
        'width': 'AC',
    },
}

CELL_REFERENCE_PATTERN = re.compile(r'^(\$?[A-Z]{1,3})(\$?)(\d+)$')

LINE_I_KEY_REPLACEMENTS = (
    ('Line-3', 'Line-1'),
    ('Line-4', 'Line-2'),
    ('Auto trimming - 3', 'Auto trimming - 1'),
    ('Auto trimming - 4', 'Auto trimming - 2'),
)

LINE_I_LINE_SELECTION_CELL_MAPPINGS = {
    '2-4': {'4h': 'M16', '8h': 'Y16'}, '2-5': {'4h': 'M16', '8h': 'Y16'}, '2-6': {'4h': 'M16', '8h': 'Y16'},
    '3-7': {'4h': 'M27', '8h': 'Y27'}, '3-8': {'4h': 'M27', '8h': 'Y27'}, '3-9': {'4h': 'M27', '8h': 'Y27'},
    '9-6': {'4h': 'M226', '8h': 'Y226'}, '9-7': {'4h': 'M226', '8h': 'Y226'}, '9-8': {'4h': 'M226', '8h': 'Y226'},
    '10-4': {'4h': 'M233', '8h': 'Y233'}, '10-5': {'4h': 'M233', '8h': 'Y233'}, '10-6': {'4h': 'M233', '8h': 'Y233'},
    '11-3': {'4h': 'M237', '8h': 'Y237'}, '11-4': {'4h': 'M237', '8h': 'Y237'}, '11-5': {'4h': 'M237', '8h': 'Y237'},
}

LINE_I_GROUPED_LINE_SELECTION_CELL_MAPPINGS = {
    '12-1': {'2h': 'J245', '4h': 'P245', '6h': 'V245', '8h': 'AB245'},
    '12-2': {'2h': 'J245', '4h': 'P245', '6h': 'V245', '8h': 'AB245'},
    '16-1': {'2h': 'J284', '4h': 'P284', '6h': 'V284', '8h': 'AB284'},
    '23-1': {'2h': 'J366', '4h': 'P366', '6h': 'V366', '8h': 'AB366'},
    '27-1': {'2h': 'J409', '4h': 'P409', '6h': 'V409', '8h': 'AB409'},
    '29-1': {'2h': 'J415', '4h': 'P415', '6h': 'V415', '8h': 'AB415'},
}

def build_stringer_unit_cell_mapping(stringer_rows, fields, columns):
    return {
        f'Stringer-{stringer_number}': {
            'unitA': {
                field_name: f'{column}{unit_a_row}'
                for field_name, column in zip(fields, columns)
            },
            'unitB': {
                field_name: f'{column}{unit_b_row}'
                for field_name, column in zip(fields, columns)
            },
        }
        for stringer_number, unit_a_row, unit_b_row in stringer_rows
    }

def build_line_i_machine_temp_mapping():
    columns = ('K', 'M', 'O', 'Q', 'S', 'U', 'V', 'X', 'Y', 'Z', 'AA', 'AB', 'AC', 'AD')
    return build_stringer_unit_cell_mapping(
        LINE_I_STRINGER_SETUP_ROWS,
        LINE_I_MACHINE_TEMP_FIELDS,
        columns,
    )

def build_line_i_light_intensity_mapping():
    columns = tuple(get_column_letter(column_index) for column_index in range(9, 31))
    return build_stringer_unit_cell_mapping(
        LINE_I_LIGHT_INTENSITY_ROWS,
        LINE_I_LIGHT_INTENSITY_FIELDS,
        columns,
    )

def build_line_i_auto_bussing_soldering_time_mapping():
    return {
        line: {
            f'{position}_tca_{tca_number}': f'{column}{position_rows[position]}'
            for tca_number, column in enumerate(LINE_I_AUTO_BUSSING_SOLDERING_TIME_COLUMNS, start=1)
            for position in ('front', 'middle', 'back')
        }
        for line, position_rows in LINE_I_AUTO_BUSSING_SOLDERING_TIME_ROWS.items()
    }

def build_line_i_stringer_stage_mapping():
    stringer_layout = (
        (1, 42, 43, 45, 47, 'G', 'K', ('H', 'J', 'L', 'N')),
        (2, 42, 43, 45, 47, 'O', 'S', ('P', 'R', 'T', 'V')),
        (3, 42, 43, 45, 47, 'W', 'AA', ('X', 'Z', 'AB', 'AD')),
        (4, 50, 51, 53, 55, 'G', 'K', ('H', 'J', 'L', 'N')),
        (5, 50, 51, 53, 55, 'O', 'S', ('P', 'R', 'T', 'V')),
        (6, 50, 51, 53, 55, 'W', 'AA', ('X', 'Z', 'AB', 'AD')),
    )
    laser_power = {}
    cell_appearance = {}
    cell_width = {}
    groove_length = {}
    peel_strength = {}
    ribbon_flatten = {}
    string_length = {}
    cell_gap = {}
    el_inspection = {}

    for stringer_number, laser_row, appearance_row, width_row, groove_row, unit_a_column, unit_b_column, measurement_columns in stringer_layout:
        stringer_key = f'Stringer-{stringer_number}'
        laser_power[stringer_key] = {
            'Unit A': f'{unit_a_column}{laser_row}',
            'Unit B': f'{unit_b_column}{laser_row}',
        }
        cell_appearance[stringer_key] = {
            'Unit A': f'{unit_a_column}{appearance_row}',
            'Unit B': f'{unit_b_column}{appearance_row}',
        }
        upper_a_column, lower_a_column, upper_b_column, lower_b_column = measurement_columns
        cell_width[stringer_key] = {
            'Upper-A-L': f'{upper_a_column}{width_row}',
            'Upper-A-R': f'{upper_a_column}{width_row + 1}',
            'Lower-A-L': f'{lower_a_column}{width_row}',
            'Lower-A-R': f'{lower_a_column}{width_row + 1}',
            'Upper-B-L': f'{upper_b_column}{width_row}',
            'Upper-B-R': f'{upper_b_column}{width_row + 1}',
            'Lower-B-L': f'{lower_b_column}{width_row}',
            'Lower-B-R': f'{lower_b_column}{width_row + 1}',
        }
        groove_length[stringer_key] = {
            'Unit A - Upper Half': f'{upper_a_column}{groove_row}',
            'Unit A - Lower Half': f'{lower_a_column}{groove_row}',
            'Unit B - Upper Half': f'{upper_b_column}{groove_row}',
            'Unit B - Lower Half': f'{lower_b_column}{groove_row}',
        }

        peel_header_row = 84 + ((stringer_number - 1) * 8)
        peel_strength[stringer_key] = {
            'frontUnit': f'H{peel_header_row}',
            'backUnit': f'H{peel_header_row + 2}',
            'frontSide': {
                str(position): f'{get_column_letter(position + 10)}{peel_header_row + 1}'
                for position in range(1, 21)
            },
            'backSide': {
                str(position): f'{get_column_letter(position + 10)}{peel_header_row + 3}'
                for position in range(1, 21)
            },
        }
        ribbon_flatten[stringer_key] = f'G{peel_header_row + 4}'
        string_length[stringer_key] = {
            '4 hours': f'G{peel_header_row + 5}',
            '8 hours': f'S{peel_header_row + 5}',
        }
        cell_gap[stringer_key] = {
            '4 hours': f'G{peel_header_row + 6}',
            '8 hours': f'S{peel_header_row + 6}',
        }
        el_inspection[stringer_key] = {
            '4 hours': f'G{peel_header_row + 7}',
            '8 hours': f'S{peel_header_row + 7}',
        }

    return {
        '5-4-laser-power': laser_power,
        '5-5-cell-appearance': cell_appearance,
        '5-6-cell-width': cell_width,
        '5-7-groove-length': groove_length,
        '5-9-machine-temp-setup': build_line_i_machine_temp_mapping(),
        '5-10-light-intensity-time': build_line_i_light_intensity_mapping(),
        '5-11-peel-strength': peel_strength,
        '5-12-ribbon-flatten': ribbon_flatten,
        '5-13-string-length': string_length,
        '5-14-cell-to-cell-gap': cell_gap,
        '5-15-el-inspection': el_inspection,
    }

def build_line_i_auto_bussing_mapping():
    bus_ribbon_status = {}
    cooling_temperature = {}
    soldering_iron_temperature = {}
    soldering_trip_calibration = {}
    soldering_traces = {}
    bus_bar_cut_length = {}
    string_alignment = {}
    patch_measurements = {}
    peel_strength = {}

    for line, start_row in LINE_I_AUTO_BUSSING_BLOCK_ROWS.items():
        bus_ribbon_status.update({
            f'{line}-Supplier': f'I{start_row}',
            f'{line}-Width Top & Bottom': f'O{start_row}',
            f'{line}-Width Middle': f'P{start_row}',
            f'{line}-Thickness Top & Bottom': f'U{start_row}',
            f'{line}-Thickness Middle': f'V{start_row}',
            f'{line}-Expiry Date Top & Bottom': f'AA{start_row}',
            f'{line}-Expiry Date Middle': f'AB{start_row}',
        })
        cooling_temperature.update({
            f'{line}-left': f'G{start_row + 4}',
            f'{line}-right': f'S{start_row + 4}',
        })
        soldering_iron_temperature.update({
            f'{line}-4hrs': f'G{start_row + 5}',
            f'{line}-8hrs': f'S{start_row + 5}',
        })
        soldering_trip_calibration[line] = f'G{start_row + 6}'
        soldering_traces.update({
            f'{line}-Top': f'I{start_row + 7}',
            f'{line}-Middle': f'Q{start_row + 7}',
            f'{line}-Bottom': f'Y{start_row + 7}',
        })
        bus_bar_cut_length.update({
            f'{line}-I': f'J{start_row + 8}',
            f'{line}-Small L': f'V{start_row + 9}',
            f'{line}-Big L': f'J{start_row + 9}',
            f'{line}-Terminal': f'V{start_row + 8}',
        })
        string_alignment.update({
            f'{line}-4hrs': f'G{start_row + 10}',
            f'{line}-8hrs': f'S{start_row + 10}',
        })
        patch_row = start_row + 11
        patch_measurements[line] = {
            sample_key: {
                field_name: f'{column}{patch_row}'
                for field_name, column in field_columns.items()
            }
            for sample_key, field_columns in AUTO_BUSSING_PATCH_VALUE_COLUMNS.items()
        }
        peel_header_row = start_row + 12
        peel_strength.update({
            f'{line}-Line': f'I{peel_header_row}',
            f'{line}-Position': f'O{peel_header_row}',
            f'{line}-Side': f'V{peel_header_row}',
            **{
                f'{line}-Pos{position}': f'{get_column_letter(((position - 1) % 20) + 11)}{peel_header_row + 1 + ((position - 1) // 20)}'
                for position in range(1, 41)
            },
        })

    return {
        '7-1': bus_ribbon_status,
        '7-2': build_line_i_auto_bussing_soldering_time_mapping(),
        '7-3': cooling_temperature,
        '7-4': soldering_iron_temperature,
        '7-5': soldering_trip_calibration,
        '7-6': soldering_traces,
        '7-7': bus_bar_cut_length,
        '7-8': string_alignment,
        AUTO_BUSSING_PATCH_PARAMETER_ID: patch_measurements,
        '7-9': peel_strength,
    }

def replace_mapping_key(key, replacements):
    if not isinstance(key, str):
        return key
    updated_key = key
    for old_value, new_value in replacements:
        updated_key = updated_key.replace(old_value, new_value)
    return updated_key

def offset_cell_reference(cell_ref, row_offset):
    if not isinstance(cell_ref, str):
        return cell_ref
    match = CELL_REFERENCE_PATTERN.fullmatch(cell_ref)
    if not match:
        return cell_ref
    column, absolute_row_marker, row_text = match.groups()
    return f'{column}{absolute_row_marker}{int(row_text) + row_offset}'

def remap_cell_mapping(value, row_offset=0, key_replacements=()):
    if isinstance(value, dict):
        return {
            replace_mapping_key(key, key_replacements): remap_cell_mapping(item, row_offset, key_replacements)
            for key, item in value.items()
        }
    if isinstance(value, list):
        return [remap_cell_mapping(item, row_offset, key_replacements) for item in value]
    if isinstance(value, tuple):
        return tuple(remap_cell_mapping(item, row_offset, key_replacements) for item in value)
    return offset_cell_reference(value, row_offset)

def apply_line_i_mapping_overrides(observation_cell_mapping):
    observation_cell_mapping.update(build_line_i_stringer_stage_mapping())
    observation_cell_mapping.update(build_line_i_auto_bussing_mapping())

    for param_id in list(observation_cell_mapping):
        try:
            stage_id = int(param_id.split('-', 1)[0])
        except (TypeError, ValueError):
            continue

        if stage_id == 8:
            observation_cell_mapping[param_id] = remap_cell_mapping(
                observation_cell_mapping[param_id],
                row_offset=21,
                key_replacements=LINE_I_KEY_REPLACEMENTS,
            )
        elif stage_id == 9:
            if param_id == '9-9':
                observation_cell_mapping.pop(param_id, None)
            else:
                observation_cell_mapping[param_id] = remap_cell_mapping(
                    observation_cell_mapping[param_id],
                    row_offset=35,
                    key_replacements=LINE_I_KEY_REPLACEMENTS,
                )
        elif 10 <= stage_id <= 31:
            observation_cell_mapping[param_id] = remap_cell_mapping(
                observation_cell_mapping[param_id],
                row_offset=34,
                key_replacements=LINE_I_KEY_REPLACEMENTS,
            )

    for index in range(1, 5):
        line_ii_param_id = f'14-{index}-laminator{index + 4}'
        line_i_param_id = f'14-{index}-laminator{index}'
        if line_ii_param_id in observation_cell_mapping:
            observation_cell_mapping[line_i_param_id] = observation_cell_mapping.pop(line_ii_param_id)

    # Line-I has a third Auto Taping & Layup block that cannot be derived from
    # the two-block Line-II template mapping.
    observation_cell_mapping.update(build_line_i_auto_taping_mapping())

    return observation_cell_mapping

LINE_I_AUTO_TAPING_BLOCK_ROWS = {
    'Line-1': 178,
    'Line-2': 192,
    'Line-3': 206,
}
LINE_I_AUTO_TAPING_TIME_ROW_OFFSETS = {
    '8-1': 0,
    '8-2': 1,
    '8-3': 2,
    '8-4': 3,
    '8-5': 4,
    '8-9': 6,
    '8-10': 7,
    '8-11': 8,
    '8-12': 9,
    '8-13': 10,
    '8-14': 11,
    '8-15': 12,
    '8-16': 13,
}

def build_line_i_auto_taping_mapping():
    auto_taping_mapping = {param_id: {} for param_id in LINE_I_AUTO_TAPING_TIME_ROW_OFFSETS}
    cell_fixing_tape_mapping = {}

    for line, start_row in LINE_I_AUTO_TAPING_BLOCK_ROWS.items():
        for param_id, row_offset in LINE_I_AUTO_TAPING_TIME_ROW_OFFSETS.items():
            auto_taping_mapping[param_id].update({
                f'{line}-4hrs': f'G{start_row + row_offset}',
                f'{line}-8hrs': f'S{start_row + row_offset}',
            })
        cell_fixing_tape_mapping.update({
            f'{line}-Supplier': f'I{start_row + 5}',
            f'{line}-Type': f'P{start_row + 5}',
            f'{line}-Quantity': f'Y{start_row + 5}',
        })

    auto_taping_mapping['8-6'] = cell_fixing_tape_mapping
    return auto_taping_mapping
