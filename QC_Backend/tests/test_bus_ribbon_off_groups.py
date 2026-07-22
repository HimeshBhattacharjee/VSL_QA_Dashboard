import unittest

from fastapi import HTTPException
from openpyxl import Workbook

from generators.BusRibbonPullStrengthReportGenerator import write_machine_rows
from routes.bus_ribbon_pull_strength_route import normalize_entry_payload
from services.bus_ribbon_group_service import is_bussing_group_off


def build_entry(machine_override=None):
    machine = {"isOff": False, "position": "1-TOP", "strengths": ["2"] * 32}
    machine.update(machine_override or {})
    return {
        "date": "2026-07-21",
        "testingDate": "2026-07-21",
        "shift": "A",
        "line": "FAB-II Line-I",
        "shiftDetails": {},
        "bussingData": {
            "autoBussing1": machine,
            "autoBussing2": {"isOff": False, "position": "", "strengths": []},
            "autoBussing3": {"isOff": False, "position": "", "strengths": []},
        },
    }


class BusRibbonOffGroupTests(unittest.TestCase):
    def test_valid_off_payload_is_type_safe_and_has_no_average(self):
        normalized = normalize_entry_payload(build_entry({"isOff": True, "position": "", "strengths": []}))
        machine = normalized["bussingData"]["autoBussing1"]
        self.assertTrue(machine["isOff"])
        self.assertEqual(machine["strengths"], [""] * 32)
        self.assertEqual(normalized["averages"]["autoBussing1"], {"average1": "", "average2": ""})

    def test_off_payload_with_measurements_is_rejected(self):
        with self.assertRaises(HTTPException) as raised:
            normalize_entry_payload(build_entry({"isOff": True, "position": "", "strengths": ["2"]}))
        self.assertEqual(raised.exception.status_code, 400)
        self.assertIn("while OFF", raised.exception.detail)

    def test_normal_numeric_payload_regression(self):
        normalized = normalize_entry_payload(build_entry())
        self.assertFalse(normalized["bussingData"]["autoBussing1"]["isOff"])
        self.assertEqual(normalized["averages"]["autoBussing1"], {"average1": "2.00", "average2": "2.00"})

    def test_legacy_position_off_is_loaded_as_off(self):
        self.assertTrue(is_bussing_group_off({"position": "OFF", "strengths": []}))

    def test_excel_writes_off_to_position_all_readings_and_averages(self):
        workbook = Workbook()
        worksheet = workbook.active
        entry = build_entry({"isOff": True, "position": "", "strengths": []})
        entry["averages"] = {"autoBussing1": {"average1": "", "average2": ""}}
        write_machine_rows(worksheet, entry, "autoBussing1", 7)

        self.assertEqual(worksheet["H7"].value, "OFF")
        for row in (7, 8):
            for column in range(9, 25):
                self.assertEqual(worksheet.cell(row=row, column=column).value, "OFF")
            self.assertEqual(worksheet.cell(row=row, column=25).value, "OFF")


if __name__ == "__main__":
    unittest.main()
