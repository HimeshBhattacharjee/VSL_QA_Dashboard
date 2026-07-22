from openpyxl import Workbook

from generators.FrameSealantWtReportGenerator import fill_shift_data
from generators.JBSealantWeightReportGenerator import fill_jb_data_in_sheet
from generators.PeelStrengthBusRibbonJBSolderingReportGenerator import write_line_row
from generators.PottingRatioReportGenerator import fill_potting_data_in_sheet
from generators.SSHReportGenerator import fill_line_row


def sheet():
    return Workbook().active


def test_frame_off_is_scoped_and_maps_line_ii_to_3_4():
    ws = sheet()
    fill_shift_data(ws, {"shift": "A", "lineGroup": "Line-II", "lines": {"1": {"status": "OFF"}, "2": {"status": "ON", "po": "PO-2", "length": {}, "width": {}}}}, "2026-07-21", 7)
    assert ws["A7"].value == "2026-07-21" and ws["B7"].value == "Shift A"
    assert ws["A8"].value == "2026-07-21" and ws["B8"].value == "Shift A"
    assert ws["C7"].value == "3" and ws["D7"].value == "OFF" and ws["Q8"].value == "OFF"
    assert ws["C9"].value == "4" and ws["D9"].value == "PO-2"


def test_jb_off_fills_only_its_three_rows():
    ws = sheet()
    fill_jb_data_in_sheet(ws, [{"shift": "A", "lineGroup": "Line-I", "lines": {"1": {"status": "OFF"}, "2": {"status": "ON", "po": "LIVE"}}}], "21.07.2026")
    assert all(ws[f"D{row}"].value == "OFF" for row in range(5, 8))
    assert ws["D8"].value == "LIVE"


def test_potting_off_fills_correct_physical_line_only():
    ws = sheet()
    fill_potting_data_in_sheet(ws, [{"shift": "B", "lineGroup": "Line-II", "lines": {"1": {"status": "ON", "po": "LIVE"}, "2": {"status": "OFF"}}}], "21.07.2026")
    assert ws["D6"].value == "3" and ws["E6"].value == "LIVE"
    assert ws["D7"].value == "4" and all(ws[f"{column}7"].value == "OFF" for column in "EFGHIJK")


def test_ssh_off_row_contains_only_off_values():
    ws = sheet()
    fill_line_row(ws, 10, {"line": "2", "status": "OFF", "result": "10"})
    assert ws["F10"].value == "2"
    assert all(ws[f"{column}10"].value == "OFF" for column in "GHIJKLM")


def test_peel_off_row_does_not_affect_other_row():
    ws = sheet()
    write_line_row(ws, 7, "21.07.2026", "C", "Line - 3", {"status": "OFF", "average": 0})
    write_line_row(ws, 8, "21.07.2026", "C", "Line - 4", {"status": "ON", "po": "LIVE", "plusVe1": "2"})
    assert all(ws[f"{column}7"].value == "OFF" for column in "EFGHIJKLMNOP")
    assert ws["E8"].value == "LIVE" and ws["I8"].value == 2
