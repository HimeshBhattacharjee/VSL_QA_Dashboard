import pytest
from openpyxl import Workbook

from generators.SSHReportGenerator import fill_ssh_test_data
from report_context import FAB_LINE_I, FAB_LINE_II, logical_physical_line_pairs, normalize_fab_line


def entry(line_group, first, second):
    return {
        "date": "2026-07-22", "shift": "A", "lineGroup": line_group,
        "po": "PO-SENTINEL",
        "lines": {
            "1": {"result": first, "sealantSupplier": "slot-one"},
            "2": {"result": second, "sealantSupplier": "slot-two"},
        },
    }


@pytest.mark.parametrize(("context", "labels"), [("Line-I", ("1", "2")), ("Line-II", ("3", "4"))])
def test_export_labels_and_sentinel_measurements_follow_fab_line(context, labels):
    ws = Workbook().active
    fill_ssh_test_data(ws, [entry(context, 31.1, 42.2)], context)
    assert (ws["F6"].value, ws["F7"].value) == labels
    assert (ws["G6"].value, ws["G7"].value) == ("slot-one", "slot-two")
    assert (ws["K6"].value, ws["K7"].value) == (31.1, 42.2)


@pytest.mark.parametrize(("legacy", "canonical", "pairs"), [
    ("line i", FAB_LINE_I, (("1", "1"), ("2", "2"))),
    ("FAB II LINE II", FAB_LINE_II, (("1", "3"), ("2", "4"))),
    ("II", FAB_LINE_II, (("1", "3"), ("2", "4"))),
])
def test_legacy_context_normalization(legacy, canonical, pairs):
    assert normalize_fab_line(legacy) == canonical
    assert logical_physical_line_pairs(legacy) == pairs


@pytest.mark.parametrize("invalid", [None, "", "Line-III", "FAB-I Line-I"])
def test_unknown_context_is_explicit_error(invalid):
    with pytest.raises(ValueError, match="fab_line must be one of"):
        normalize_fab_line(invalid)


def test_mixed_fab_context_cannot_generate_contradictory_workbook():
    with pytest.raises(ValueError, match="exactly one valid FAB line context"):
        fill_ssh_test_data(Workbook().active, [entry("Line-I", 11, 12), entry("Line-II", 33, 44)])
